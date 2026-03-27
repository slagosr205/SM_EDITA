"""Ventana Principal del Sistema - Navegacion por contenido"""
import tkinter as tk
from tkinter import ttk, messagebox
from auth import UserSession
from styles import ElegantStyles
from dashboard import DashboardWindow
from database import get_connection, get_impresiones_count, registrar_impresion, get_contrato_completo, generar_contrato_pdf, get_config_sistema, update_config_sistema, get_monedas
import datetime
from datetime import datetime, timedelta
import random

class RoundedFrame(tk.Canvas):
    def __init__(self, parent, bg="white", radius=5, **kwargs):
        tk.Canvas.__init__(self, parent, bg=bg, highlightthickness=0, **kwargs)
        self.bg = bg
        self.radius = radius
        self.bind("<Configure>", self.draw)
        
    def draw(self, event=None):
        self.delete("border")
        w, h = self.winfo_width(), self.winfo_height()
        if w < 2 or h < 2:
            return
        r = min(self.radius, w//2, h//2)
        self.create_oval(0, 0, 2*r, 2*r, fill=self.bg, outline=self.bg, tags="border")
        self.create_oval(w-2*r, 0, w, 2*r, fill=self.bg, outline=self.bg, tags="border")
        self.create_oval(0, h-2*r, 2*r, h, fill=self.bg, outline=self.bg, tags="border")
        self.create_oval(w-2*r, h-2*r, w, h, fill=self.bg, outline=self.bg, tags="border")
        self.create_rectangle(r, 0, w-r, h, fill=self.bg, outline=self.bg, tags="border")
        self.create_rectangle(0, r, w, h-r, fill=self.bg, outline=self.bg, tags="border")

class MainWindow:
    def __init__(self, usuario_actual):
        self.usuario_actual = usuario_actual
        self.session = UserSession.get_instance()
        self.session.set_usuario(usuario_actual)
        self.S = ElegantStyles
        self.current_module = None
        self.module_frames = {}
        
        self.carrito = []
        self.pos_productos = []
        self.subtotal_label = None
        self.impuesto_label = None
        self.total_label = None
        self.total_usd_label = None
        self.cart_tree = None
        self.pos_tree = None
        
        self.root = tk.Tk()
        self.root.title("Tienda Concepto - Panel de Control")
        self.root.state("zoomed")
        self.root.configure(bg=self.S.COLORS["bg_main"])
        ElegantStyles.configure_styles(self.root)
        
        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#f1f5f9", foreground="#1e293b")
        style.map("Treeview", background=[("selected", "#3b82f6")], foreground=[("selected", "white")])
        
        self.create_widgets()
        self.show_module("dashboard")
        
    def center_window(self):
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (700)
        y = (self.root.winfo_screenheight() // 2) - (400)
        self.root.geometry(f"1400x800+{x}+{y}")
        
    def create_widgets(self):
        # Top Navbar
        navbar = tk.Frame(self.root, bg="#0f172a", height=65)
        navbar.pack(fill="x")
        navbar.pack_propagate(False)
        
        # Logo en navbar
        logo_frame = tk.Frame(navbar, bg="#0f172a")
        logo_frame.pack(side="left", padx=20, pady=10)
        
        tk.Label(logo_frame, text="TC", font=("Segoe UI", 24, "bold"),
                bg="#0f172a", fg="#14b8a6").pack(side="left")
        tk.Label(logo_frame, text="TIENDA CONCEPTO", font=("Segoe UI", 9, "bold"),
                bg="#0f172a", fg="#94a3b8").pack(side="left", padx=(10, 0))
        
        # Navigation buttons
        nav_frame = tk.Frame(navbar, bg="#0f172a")
        nav_frame.pack(side="left", padx=30)
        
        rol = self.usuario_actual["rol"]
        
        menu_items = []
        
        if rol in ["admin", "supervisor"]:
            menu_items.append(("🏠", "Dashboard", "dashboard"))
            
        if rol == "admin":
            menu_items.append(("🏢", "Proveedores", "proveedores"))
            menu_items.append(("📄", "Contratos", "contratos"))
            menu_items.append(("📦", "Productos", "productos"))
            
        menu_items.append(("📋", "Inventario", "inventario"))
        
        if rol in ["admin", "supervisor"]:
            menu_items.append(("🚚", "Surtido", "surtido"))
            
        menu_items.append(("💰", "Punto de Venta", "pos"))
        menu_items.append(("👥", "Clientes", "clientes"))
        
        if rol in ["admin", "supervisor"]:
            menu_items.append(("💵", "Liquidaciones", "liquidaciones"))
            
        if rol == "admin":
            menu_items.append(("⚙️", "Configuracion", "configuracion"))
        
        self.nav_buttons = {}
        for icon, text, module_id in menu_items:
            btn = tk.Button(nav_frame, text=f"{icon} {text}",
                          font=("Segoe UI", 11),
                          bg="#0f172a", fg="white",
                          padx=15, pady=8,
                          relief="flat", cursor="hand2",
                          command=lambda m=module_id: self.show_module(m))
            btn.pack(side="left", padx=3)
            self.nav_buttons[module_id] = btn
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#334155"))
            btn.bind("<Leave>", lambda e, b=btn, mod=module_id: b.config(bg="#0f172a" if self.current_module != mod else "#14b8a6"))
        
        # User info a la derecha
        user_frame = tk.Frame(navbar, bg="#0f172a")
        user_frame.pack(side="right", padx=20, pady=10)
        
        tk.Label(user_frame, text=f"{self.usuario_actual['nombre']}",
                font=("Segoe UI", 11, "bold"), bg="#0f172a", fg="white").pack(anchor="e")
        
        rol_colors = {"admin": "#14b8a6", "supervisor": "#3b82f6", "cajero": "#8b5cf6", "proveedor": "#f59e0b"}
        rol_bg = rol_colors.get(rol.lower(), "#8b5cf6")
        tk.Label(user_frame, text=rol.upper(), font=("Segoe UI", 9, "bold"),
                bg=rol_bg, fg="white", padx=10, pady=2).pack(anchor="e", pady=(3, 0))
        
        tk.Button(user_frame, text="Cerrar Sesion",
                 font=("Segoe UI", 9),
                 bg="#0f172a", fg="#ef4444",
                 relief="flat", cursor="hand2",
                 command=self.cerrar_sesion).pack(anchor="e", pady=(5, 0))
        
        # Main Area
        self.main_area = tk.Frame(self.root, bg=self.S.COLORS["bg_main"])
        self.main_area.pack(fill="both", expand=True)
        
        # Header
        self.header = tk.Frame(self.main_area, bg="white", height=60)
        self.header.pack(fill="x")
        self.header.pack_propagate(False)
        
        self.page_title = tk.Label(self.header, text="Dashboard",
                                  font=("Segoe UI", 18, "bold"),
                                  bg="white", fg=self.S.COLORS["text_dark"])
        self.page_title.pack(side="left", padx=25, pady=12)
        
        now = datetime.now()
        date_str = now.strftime("%d %B, %Y")
        tk.Label(self.header, text=date_str,
                font=("Segoe UI", 10),
                bg="white", fg=self.S.COLORS["text_light"]).pack(side="right", padx=25, pady=18)
        
        # Content Frame
        self.content_frame = tk.Frame(self.main_area, bg=self.S.COLORS["bg_main"])
        self.content_frame.pack(fill="both", expand=True, padx=25, pady=15)
        
    def update_nav_buttons(self):
        for module_id, btn in self.nav_buttons.items():
            if module_id == self.current_module:
                btn.config(bg="#14b8a6", fg="white")
            else:
                btn.config(bg="#0f172a", fg="white")
                
    def get_currency_info(self):
        config = get_config_sistema()
        monedas = get_monedas()
        simbolo_local = next((m["simbolo"] for m in monedas if m["codigo"] == config["moneda_local"]), "L")
        simbolo_extranjera = next((m["simbolo"] for m in monedas if m["codigo"] == config["moneda_extranjera"]), "$")
        return {
            "local": config["moneda_local"],
            "extranjera": config["moneda_extranjera"],
            "simbolo_local": simbolo_local,
            "simbolo_extranjera": simbolo_extranjera,
            "tasa": config["tasa_cambio"]
        }
    
    def format_price(self, amount, currency="local"):
        info = self.get_currency_info()
        if currency == "local":
            return f"{info['simbolo_local']} {amount:,.2f}"
        else:
            return f"{info['simbolo_extranjera']} {amount:,.2f}"
    
    def format_price_both(self, amount):
        info = self.get_currency_info()
        local = f"{info['simbolo_local']} {amount:,.2f}"
        usd = f"{info['simbolo_extranjera']} {amount / info['tasa']:,.2f}"
        return local, usd
    
    def show_module(self, module_id):
        self.current_module = module_id
        self.update_nav_buttons()
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        module_titles = {
            "dashboard": "Dashboard",
            "proveedores": "Proveedores",
            "contratos": "Contratos y Espacios",
            "productos": "Productos",
            "inventario": "Inventario",
            "surtido": "Surtido",
            "pos": "Punto de Venta",
            "clientes": "Clientes",
            "liquidaciones": "Liquidaciones",
            "configuracion": "Configuracion"
        }
        
        self.page_title.config(text=module_titles.get(module_id, module_id))
        
        if module_id == "dashboard":
            self.dashboard = DashboardWindow(self)
        elif module_id == "proveedores":
            self.show_proveedores()
        elif module_id == "contratos":
            self.show_contratos()
        elif module_id == "productos":
            self.show_productos()
        elif module_id == "inventario":
            self.show_inventario()
        elif module_id == "surtido":
            self.show_surtido()
        elif module_id == "pos":
            self.show_pos()
        elif module_id == "clientes":
            self.show_clientes()
        elif module_id == "liquidaciones":
            self.show_liquidaciones()
        elif module_id == "configuracion":
            self.show_configuracion()
            
    def show_dashboard(self):
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COALESCE(SUM(total), 0) as total FROM ventas WHERE DATE(fecha_venta) = DATE('now') AND estado = 'completada'")
        ventas_hoy = cursor.fetchone()["total"]
        
        cursor.execute("SELECT COUNT(*) as total FROM proveedores WHERE estado='activo'")
        total_proveedores = cursor.fetchone()["total"]
        
        cursor.execute("SELECT COUNT(*) as total FROM productos WHERE estado='activo'")
        total_productos = cursor.fetchone()["total"]
        
        cursor.execute("SELECT COUNT(*) as total FROM inventarios WHERE cant_actual <= cant_minima")
        alertas = cursor.fetchone()["total"]
        
        cursor.execute("SELECT COUNT(*) as total FROM clientes")
        total_clientes = cursor.fetchone()["total"]
        
        conn.close()
        
        metrics_frame = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        metrics_frame.pack(fill="x", pady=(0, 25))
        
        cards_data = [
            ("Ventas Hoy", f"S/. {ventas_hoy:,.2f}", self.S.COLORS["accent"]),
            ("Proveedores", str(total_proveedores), self.S.COLORS["success"]),
            ("Productos", str(total_productos), self.S.COLORS["purple"]),
            ("Alertas Stock", str(alertas), self.S.COLORS["danger"]),
            ("Clientes", str(total_clientes), self.S.COLORS["teal"]),
        ]
        
        for title, value, color in cards_data:
            card = tk.Frame(metrics_frame, bg=color, relief="flat")
            card.pack(side="left", fill="both", expand=True, padx=8)
            tk.Label(card, text=title, font=("Segoe UI", 11, "bold"), bg=color, fg="white").pack(pady=(15, 5), padx=15)
            tk.Label(card, text=value, font=("Segoe UI", 24, "bold"), bg=color, fg="white").pack(pady=(0, 15), padx=15)
        
        bottom_frame = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        bottom_frame.pack(fill="both", expand=True)
        
        left_panel = tk.Frame(bottom_frame, bg="white", relief="flat")
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        tk.Label(left_panel, text="Ultimas Transacciones", font=("Segoe UI", 14, "bold"), bg="white", fg=self.S.COLORS["text_dark"]).pack(anchor="w", padx=20, pady=15)
        
        columns = ("ID", "Fecha", "Cliente", "Total", "Metodo")
        tree = ttk.Treeview(left_panel, columns=columns, show="headings", height=12)
        for col in columns:
            tree.heading(col, text=col)
            widths = {"ID": 50, "Fecha": 150, "Cliente": 180, "Total": 100, "Metodo": 100}
            tree.column(col, width=widths.get(col, 100))
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True, padx=(15, 0), pady=(0, 15))
        scrollbar.pack(side="right", fill="y", padx=(0, 15), pady=(0, 15))
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT v.id_venta, v.fecha_venta, COALESCE(c.nombre, 'Consumidor Final') as cliente, v.total, v.metodo_pago
            FROM ventas v LEFT JOIN clientes c ON v.id_cliente = c.id_cliente ORDER BY v.fecha_venta DESC LIMIT 10
        """)
        for row in cursor.fetchall():
            tree.insert("", "end", values=(row["id_venta"], row["fecha_venta"], row["cliente"], f"S/. {row['total']:.2f}", row["metodo_pago"] or ""))
        conn.close()
        
        right_panel = tk.Frame(bottom_frame, bg="white", relief="flat")
        right_panel.pack(side="left", fill="both", expand=True)
        
        tk.Label(right_panel, text="Resumen de Inventario", font=("Segoe UI", 14, "bold"), bg="white", fg=self.S.COLORS["text_dark"]).pack(anchor="w", padx=20, pady=15)
        
        text_resumen = tk.Text(right_panel, width=45, height=14, font=("Segoe UI", 10), bg="white", relief="flat", bd=0)
        text_resumen.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.nombre as producto, i.cant_actual, i.cant_minima, e.codigo as espacio,
                   CASE WHEN i.cant_actual <= i.cant_minima THEN 'BAJO' ELSE 'OK' END as estado
            FROM inventarios i JOIN productos p ON i.id_producto = p.id_producto
            JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio ORDER BY i.cant_actual ASC
        """)
        inventarios = cursor.fetchall()
        total_bajo = 0
        text_resumen.insert("end", "=" * 45 + "\n  RESUMEN DE INVENTARIO\n" + "=" * 45 + "\n\n")
        for inv in inventarios:
            estado_txt = " ALERTA" if inv["estado"] == "BAJO" else ""
            text_resumen.insert("end", f"> {inv['producto'][:30]}\n")
            text_resumen.insert("end", f"  Espacio: {inv['espacio']} | Stock: {inv['cant_actual']}/{inv['cant_minima']}{estado_txt}\n\n")
            if inv["estado"] == "BAJO":
                total_bajo += 1
        text_resumen.insert("end", "-" * 45 + f"\n  Total: {len(inventarios)} | OK: {len(inventarios)-total_bajo} | Alertas: {total_bajo}\n")
        conn.close()
        
    def show_proveedores(self):
        self.create_toolbar([
            ("+ Nuevo", self.nuevo_proveedor, self.S.COLORS["success"]),
            ("Editar", self.editar_proveedor, self.S.COLORS["accent"]),
        ])
        
        search_frame = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        search_frame.pack(fill="x", pady=(0, 15))
        
        search_box = tk.Frame(search_frame, bg="white", padx=10)
        search_box.pack(side="left")
        
        tk.Label(search_box, text="Buscar:", bg="white", font=("Segoe UI", 10)).pack(side="left", padx=(10, 5))
        
        self.prov_entry = ttk.Entry(search_box, width=40, font=("Segoe UI", 11))
        self.prov_entry.pack(side="left", ipady=8)
        self.prov_entry.bind("<KeyRelease>", lambda e: self.cargar_proveedores())
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Nombre", "RUC/NIT", "Telefono", "Email", "Estado")
        self.prov_tree = ttk.Treeview(card, columns=columns, show="headings", height=18)
        
        for col in columns:
            self.prov_tree.heading(col, text=col)
            self.prov_tree.column(col, width=120 if col == "Nombre" else 100)
        self.prov_tree.column("ID", width=50)
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.prov_tree.yview)
        self.prov_tree.configure(yscrollcommand=scrollbar.set)
        self.prov_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.prov_tree.bind("<Double-1>", lambda e: self.editar_proveedor())
        self.cargar_proveedores()
        
    def show_contratos(self):
        self.create_toolbar([
            ("+ Nuevo Contrato", self.nuevo_contrato, self.S.COLORS["success"]),
            ("+ Nuevo Espacio", self.nuevo_espacio, self.S.COLORS["accent"]),
            ("Imprimir Contrato", self.imprimir_contrato, self.S.COLORS["purple"]),
        ])
        
        self.contrato_notebook = ttk.Notebook(self.content_frame)
        self.contrato_notebook.pack(fill="both", expand=True)
        
        self.tab_contratos = tk.Frame(self.contrato_notebook, bg=self.S.COLORS["bg_main"])
        self.tab_espacios = tk.Frame(self.contrato_notebook, bg=self.S.COLORS["bg_main"])
        
        self.contrato_notebook.add(self.tab_contratos, text="  Contratos  ")
        self.contrato_notebook.add(self.tab_espacios, text="  Espacios  ")
        
        self.create_contratos_tab()
        self.create_espacios_tab()
        
    def create_contratos_tab(self):
        card = tk.Frame(self.tab_contratos, bg="white")
        card.pack(fill="both", expand=True, padx=15, pady=15)
        
        columns = ("ID", "Proveedor", "Fecha Inicio", "Fecha Fin", "Comision", "Estado", "Alerta")
        self.contrato_tree = ttk.Treeview(card, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.contrato_tree.heading(col, text=col)
            if col == "Alerta":
                self.contrato_tree.column(col, width=80, anchor="center")
            elif col == "ID":
                self.contrato_tree.column(col, width=50)
            elif col == "Comision":
                self.contrato_tree.column(col, width=80)
            else:
                self.contrato_tree.column(col, width=120)
        
        self.contrato_tree.tag_configure("estado_ok", background="#e8f5e9", foreground="#2e7d32")
        self.contrato_tree.tag_configure("estado_por_vencer", background="#fff8e1", foreground="#f57f17")
        self.contrato_tree.tag_configure("estado_vencido", background="#ffebee", foreground="#c62828")
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.contrato_tree.yview)
        self.contrato_tree.configure(yscrollcommand=scrollbar.set)
        self.contrato_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cargar_contratos()
        
    def create_espacios_tab(self):
        card = tk.Frame(self.tab_espacios, bg="white")
        card.pack(fill="both", expand=True, padx=15, pady=15)
        
        columns = ("ID", "Codigo", "Descripcion", "Ubicacion", "Capacidad", "Estado")
        self.espacio_tree = ttk.Treeview(card, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.espacio_tree.heading(col, text=col)
            self.espacio_tree.column(col, width=150 if col not in ("ID", "Capacidad") else 80)
        self.espacio_tree.column("ID", width=50)
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.espacio_tree.yview)
        self.espacio_tree.configure(yscrollcommand=scrollbar.set)
        self.espacio_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cargar_espacios()
        
    def imprimir_contrato(self):
        import os
        import webbrowser
        import traceback
        
        try:
            if not hasattr(self, 'contrato_tree') or self.contrato_tree is None:
                messagebox.showerror("Error", "Vaya a la pestana 'Contratos y Espacios' primero y seleccione un contrato.")
                return
            
            selection = self.contrato_tree.selection()
            if not selection:
                messagebox.showwarning("Sin seleccion", "Seleccione un contrato de la tabla")
                return
            
            item = self.contrato_tree.item(selection[0])
            values = item["values"]
            if not values:
                messagebox.showerror("Error", "No se pudo leer el contrato")
                return
            
            id_contrato = int(values[0])
            
            impresion_count = get_impresiones_count(id_contrato)
            tipo_impresion = "ORIGINAL" if impresion_count == 0 else f"COPIA {impresion_count}"
            
            filepath = generar_contrato_pdf(id_contrato, tipo_impresion)
            
            if not filepath or not os.path.exists(filepath):
                messagebox.showerror("Error", "No se pudo generar el PDF")
                return
            
            registrar_impresion(id_contrato, tipo_impresion, self.session.usuario_actual["id_usuario"])
            messagebox.showinfo("PDF Generado", f"Contrato #{id_contrato}\nTipo: {tipo_impresion}\n\nEl PDF se abrira automaticamente.")
            webbrowser.open(f"file:///{filepath}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
            traceback.print_exc()
        
    def show_productos(self):
        self.create_toolbar([
            ("+ Nuevo Producto", self.nuevo_producto, self.S.COLORS["success"]),
            ("Editar Producto", self.editar_producto, self.S.COLORS["accent"]),
            ("Exportar Excel", self.exportar_productos_excel, self.S.COLORS["success"]),
        ])
        
        search_frame = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        search_frame.pack(fill="x", pady=(0, 15))
        
        search_box = tk.Frame(search_frame, bg="white", padx=10)
        search_box.pack(side="left")
        
        tk.Label(search_box, text="Buscar:", bg="white", font=("Segoe UI", 10)).pack(side="left", padx=(10, 5))
        
        self.prod_entry = ttk.Entry(search_box, width=40, font=("Segoe UI", 11))
        self.prod_entry.pack(side="left", ipady=8)
        self.prod_entry.bind("<KeyRelease>", lambda e: self.cargar_productos())
        
        curr_info = self.get_currency_info()
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Codigo", "Nombre", "Proveedor", f"P. Venta ({curr_info['local']})", "Impuesto", "Stock", "Estado")
        self.prod_tree = ttk.Treeview(card, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.prod_tree.heading(col, text=col)
            widths = {"ID": 50, "Codigo": 120, "Nombre": 220, "Proveedor": 150, f"P. Venta ({curr_info['local']})": 100, "Impuesto": 70, "Stock": 70, "Estado": 80}
            self.prod_tree.column(col, width=widths.get(col, 100))
        
        self.prod_tree.tag_configure("stock_bajo", background="#ffebee")
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.prod_tree.yview)
        self.prod_tree.configure(yscrollcommand=scrollbar.set)
        self.prod_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.prod_tree.bind("<<TreeviewSelect>>", self.mostrar_stock_preview)
        self.prod_tree.bind("<Double-1>", lambda e: self.ver_detalle_stock())
        self.cargar_productos()
        
        self.stock_preview = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        self.stock_preview.pack(fill="x", pady=(15, 0))
        
        preview_card = tk.Frame(self.stock_preview, bg="white", relief="flat")
        preview_card.pack(fill="x")
        
        header_preview = tk.Frame(preview_card, bg=self.S.COLORS["primary"])
        header_preview.pack(fill="x")
        
        title_frame = tk.Frame(header_preview, bg=self.S.COLORS["primary"])
        title_frame.pack(side="left", fill="both", expand=True, padx=15, pady=10)
        
        tk.Label(title_frame, text="Vista rapida de Stock", font=("Segoe UI", 12, "bold"), 
                bg=self.S.COLORS["primary"], fg="white").pack(anchor="w")
        
        self.preview_producto_label = tk.Label(title_frame, text="Seleccione un producto", 
                font=("Segoe UI", 10), bg=self.S.COLORS["primary"], fg="#bdc3c7")
        self.preview_producto_label.pack(anchor="w", pady=(2, 0))
        
        tk.Button(header_preview, text="Ver Detalle", font=("Segoe UI", 10, "bold"),
                 bg=self.S.COLORS["accent"], fg="white", relief="flat",
                 cursor="hand2", padx=20, pady=8, command=self.ver_detalle_stock).pack(side="right", padx=15, pady=10)
        
        content_preview = tk.Frame(preview_card, bg="white")
        content_preview.pack(fill="x", padx=20, pady=15)
        
        self.preview_total_label = tk.Label(content_preview, text="Stock Total: 0", 
                font=("Segoe UI", 14, "bold"), bg="white")
        self.preview_total_label.pack(side="left")
        
        self.preview_info_label = tk.Label(content_preview, text="", 
                font=("Segoe UI", 10), bg="white", fg="gray")
        self.preview_info_label.pack(side="left", padx=30)
        
    def mostrar_stock_preview(self, event=None):
        selection = self.prod_tree.selection()
        if not selection:
            self.preview_producto_label.config(text="Seleccione un producto")
            self.preview_total_label.config(text="Stock Total: 0")
            self.preview_info_label.config(text="")
            return
        
        item = self.prod_tree.item(selection[0])
        values = item["values"]
        
        id_producto = values[0]
        nombre = values[2]
        stock_total = values[6]
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) as num_espacios, SUM(cant_actual) as total
            FROM inventarios WHERE id_producto = ?
        """, (id_producto,))
        result = cursor.fetchone()
        conn.close()
        
        self.preview_producto_label.config(text=nombre)
        result_dict = dict(result)
        self.preview_total_label.config(text=f"Stock Total: {result_dict.get('total') or 0}")
        
        total = result_dict.get('total') or 0
        num_espacios = result_dict.get('num_espacios') or 0
        
        if total == 0:
            self.preview_total_label.config(fg="#e74c3c")
            self.preview_info_label.config(text="Sin stock", fg="#e74c3c")
        elif total <= 5:
            self.preview_total_label.config(fg="#f39c12")
            self.preview_info_label.config(text=f"Stock bajo - {num_espacios} espacios", fg="#f39c12")
        else:
            self.preview_total_label.config(fg="#27ae60")
            self.preview_info_label.config(text=f"OK - {num_espacios} espacios", fg="#27ae60")
        
    def ver_detalle_stock(self):
        selection = self.prod_tree.selection()
        if not selection:
            messagebox.showwarning("Sin seleccion", "Seleccione un producto")
            return
        
        item = self.prod_tree.item(selection[0])
        values = item["values"]
        
        id_producto = values[0]
        nombre = values[2]
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.*, e.codigo, e.descripcion, e.ubicacion
            FROM inventarios i
            JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio
            WHERE i.id_producto = ?
        """, (id_producto,))
        inventarios = cursor.fetchall()
        
        cursor.execute("SELECT SUM(cant_actual) as total FROM inventarios WHERE id_producto = ?", (id_producto,))
        total = cursor.fetchone()["total"] or 0
        conn.close()
        
        modal = tk.Toplevel(self.root)
        modal.title("Detalle de Stock")
        modal.geometry("600x450")
        modal.configure(bg=self.S.COLORS["bg_main"])
        modal.transient(self.root)
        modal.grab_set()
        
        header = tk.Frame(modal, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        
        title_box = tk.Frame(header, bg=self.S.COLORS["primary"])
        title_box.pack(fill="both", expand=True, padx=20, pady=15)
        
        tk.Label(title_box, text=f"Stock: {nombre}", font=("Segoe UI", 14, "bold"),
                bg=self.S.COLORS["primary"], fg="white").pack(anchor="w")
        
        info_frame = tk.Frame(header, bg=self.S.COLORS["primary"])
        info_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        total_color = "#27ae60" if total > 5 else ("#f39c12" if total > 0 else "#e74c3c")
        
        tk.Label(info_frame, text=f"Stock Total: {total}", font=("Segoe UI", 16, "bold"),
                bg=self.S.COLORS["primary"], fg=total_color).pack(anchor="w")
        
        content = tk.Frame(modal, bg="white")
        content.pack(fill="both", expand=True, padx=20, pady=20)
        
        tk.Label(content, text="Distribucion por Espacio", font=("Segoe UI", 11, "bold"),
                bg="white", anchor="w").pack(anchor="w", pady=(0, 10))
        
        tree_frame = tk.Frame(content)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("Espacio", "Ubicacion", "Stock Actual", "Stock Min", "Stock Max", "Estado")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        
        for col in columns:
            tree.heading(col, text=col)
            widths = {"Espacio": 100, "Ubicacion": 150, "Stock Actual": 90, "Stock Min": 80, "Stock Max": 80, "Estado": 80}
            tree.column(col, width=widths.get(col, 100), anchor="center")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        if not inventarios:
            tree.insert("", "end", values=("-", "-", "-", "-", "-", "Sin asignar"))
        else:
            for inv in inventarios:
                cant = inv["cant_actual"]
                minimo = inv["cant_minima"]
                estado = "BAJO" if cant <= minimo else "OK"
                
                if estado == "BAJO":
                    tree.insert("", "end", tags=("alerta",), values=(
                        inv["codigo"], inv["ubicacion"] or "-", cant, minimo, inv["cant_maxima"], estado
                    ))
                else:
                    tree.insert("", "end", values=(
                        inv["codigo"], inv["ubicacion"] or "-", cant, minimo, inv["cant_maxima"], estado
                    ))
        
        tree.tag_configure("alerta", background="#ffebee")
        
        footer = tk.Frame(modal, bg=self.S.COLORS["bg_main"])
        footer.pack(fill="x", pady=15)
        
        tk.Button(footer, text="Cerrar", font=("Segoe UI", 11, "bold"),
                 bg=self.S.COLORS["secondary"], fg="white", relief="flat",
                 cursor="hand2", padx=30, pady=8, command=modal.destroy).pack()
        
    def show_inventario(self):
        self.create_toolbar([
            ("+ Asignar Producto", self.asignar_inventario, self.S.COLORS["success"]),
            ("Ajuste de Stock", self.ajustar_stock, self.S.COLORS["warning"]),
            ("Reposición", self.reposicion_inventario, "#7c3aed"),
        ])
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Producto", "Espacio", "Actual", "Min", "Max", "Estado")
        self.inv_tree = ttk.Treeview(card, columns=columns, show="headings", height=18)
        
        for col in columns:
            self.inv_tree.heading(col, text=col)
            widths = {"ID": 50, "Producto": 280, "Espacio": 100, "Actual": 70, "Min": 60, "Max": 60, "Estado": 80}
            self.inv_tree.column(col, width=widths.get(col, 100))
        
        self.inv_tree.tag_configure("alerta", background="#ffcccc")
        self.inv_tree.tag_configure("ok", background="#e8f5e9")
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.inv_tree.yview)
        self.inv_tree.configure(yscrollcommand=scrollbar.set)
        self.inv_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cargar_inventario()
        
    def show_surtido(self):
        self.create_toolbar([
            ("Generar Planes", self.generar_planes, self.S.COLORS["success"]),
        ])
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Producto", "Espacio", "Cant. Solicitada", "Fecha", "Estado")
        self.surtido_tree = ttk.Treeview(card, columns=columns, show="headings", height=18)
        
        for col in columns:
            self.surtido_tree.heading(col, text=col)
            widths = {"ID": 50, "Producto": 300, "Espacio": 100, "Cant. Solicitada": 120, "Fecha": 120, "Estado": 100}
            self.surtido_tree.column(col, width=widths.get(col, 100))
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.surtido_tree.yview)
        self.surtido_tree.configure(yscrollcommand=scrollbar.set)
        self.surtido_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cargar_surtido()
        
    def show_pos(self):
        self.create_pos_screen()
        
    def show_clientes(self):
        self.create_toolbar([
            ("+ Nuevo Cliente", self.nuevo_cliente, self.S.COLORS["success"]),
            ("Editar", self.editar_cliente, self.S.COLORS["accent"]),
        ])
        
        search_frame = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        search_frame.pack(fill="x", pady=(0, 15))
        
        search_box = tk.Frame(search_frame, bg="white", padx=10)
        search_box.pack(side="left")
        
        tk.Label(search_box, text="Buscar:", bg="white", font=("Segoe UI", 10)).pack(side="left", padx=(10, 5))
        
        self.cliente_entry = ttk.Entry(search_box, width=40, font=("Segoe UI", 11))
        self.cliente_entry.pack(side="left", ipady=8)
        self.cliente_entry.bind("<KeyRelease>", lambda e: self.cargar_clientes())
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Nombre", "Apellido", "Documento", "Telefono", "Email")
        self.cliente_tree = ttk.Treeview(card, columns=columns, show="headings", height=18)
        
        for col in columns:
            self.cliente_tree.heading(col, text=col)
            widths = {"ID": 50, "Nombre": 150, "Apellido": 150, "Documento": 110, "Telefono": 110, "Email": 180}
            self.cliente_tree.column(col, width=widths.get(col, 100))
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.cliente_tree.yview)
        self.cliente_tree.configure(yscrollcommand=scrollbar.set)
        self.cliente_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cliente_tree.bind("<Double-1>", lambda e: self.editar_cliente())
        self.cargar_clientes()
        
    def show_liquidaciones(self):
        self.create_toolbar([
            ("Generar Liquidacion", self.generar_liquidacion, self.S.COLORS["success"]),
        ])
        
        card = tk.Frame(self.content_frame, bg="white", relief="flat")
        card.pack(fill="both", expand=True)
        
        columns = ("ID", "Proveedor", "Periodo", "Total Ventas", "Comision", "Monto Pagar", "Estado")
        self.liq_tree = ttk.Treeview(card, columns=columns, show="headings", height=18)
        
        for col in columns:
            self.liq_tree.heading(col, text=col)
            widths = {"ID": 50, "Proveedor": 180, "Periodo": 180, "Total Ventas": 110, "Comision": 90, "Monto Pagar": 110, "Estado": 90}
            self.liq_tree.column(col, width=widths.get(col, 100))
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=self.liq_tree.yview)
        self.liq_tree.configure(yscrollcommand=scrollbar.set)
        self.liq_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        self.cargar_liquidaciones()
        
    def create_toolbar(self, buttons):
        toolbar = tk.Frame(self.content_frame, bg="white", height=50)
        toolbar.pack(fill="x", pady=(0, 15))
        toolbar.pack_propagate(False)
        
        btn_frame = tk.Frame(toolbar, bg="white")
        btn_frame.pack(side="right", padx=15, pady=10)
        
        for text, cmd, color in buttons:
            tk.Button(btn_frame, text=text, font=("Segoe UI", 10, "bold"),
                     bg=color, fg="white", relief="flat", cursor="hand2",
                     padx=15, pady=5, command=cmd).pack(side="left", padx=5)
        
    def create_pos_screen(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        self.content_frame.configure(bg="#f1f5f9")
        
        config = get_config_sistema()
        curr = self.get_currency_info()
        self.carrito = []
        
        main_frame = tk.Frame(self.content_frame, bg="#f1f5f9")
        main_frame.pack(fill="both", expand=True)
        
        header_bar = tk.Frame(main_frame, bg="#1e293b", height=50)
        header_bar.pack(fill="x")
        header_bar.pack_propagate(False)
        
        tk.Label(header_bar, text="PUNTO DE VENTA", font=("Segoe UI", 14, "bold"),
                bg="#1e293b", fg="white").pack(side="left", padx=15, pady=8)
        
        btn_style = {"font": ("Segoe UI", 10, "bold"), "relief": "flat", "cursor": "hand2", "padx": 15, "pady": 5}
        tk.Button(header_bar, text="Historial", bg="#3b82f6", fg="white", 
                 command=self.mostrar_historial_ventas, **btn_style).pack(side="right", padx=5, pady=8)
        tk.Button(header_bar, text="Devoluciones", bg="#dc2626", fg="white", 
                 command=self.gestionar_devolucion, **btn_style).pack(side="right", padx=5, pady=8)
        
        cobrar_style = {"font": ("Segoe UI", 12, "bold"), "relief": "flat", "cursor": "hand2", "padx": 20, "pady": 8}
        tk.Button(header_bar, text="COBRAR AHORA", bg="#16a34a", fg="white", 
                 command=self.procesar_venta, **cobrar_style).pack(side="right", padx=15, pady=6)
        
        content = tk.Frame(main_frame, bg="#f1f5f9")
        content.pack(fill="both", expand=True, pady=10)
        
        left_col = tk.Frame(content, bg="#e2e8f0", relief="flat")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 15))
        
        left_inner = tk.Frame(left_col, bg="white", relief="flat")
        left_inner.pack(fill="both", expand=True, padx=3, pady=3)
        
        right_col = tk.Frame(content, bg="#e2e8f0", relief="flat")
        right_col.pack(side="right", fill="both")
        
        right_canvas = tk.Canvas(right_col, bg="#e2e8f0", highlightthickness=0)
        right_canvas.pack(side="left", fill="both", expand=True)
        
        right_scroll = ttk.Scrollbar(right_col, orient="vertical", command=right_canvas.yview)
        right_scroll.pack(side="right", fill="y")
        right_canvas.configure(yscrollcommand=right_scroll.set)
        
        right_scroll_frame = tk.Frame(right_canvas, bg="#e2e8f0")
        right_canvas.create_window((0, 0), window=right_scroll_frame, anchor="nw")
        
        def on_frame_configure(event):
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))
        
        right_scroll_frame.bind("<Configure>", on_frame_configure)
        right_canvas.bind("<MouseWheel>", lambda e: right_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
        tk.Label(left_inner, text="CLIENTE", font=("Segoe UI", 12, "bold"),
                bg="#dcfce7", fg="#166534", anchor="w").pack(fill="x", padx=15, pady=10)
        
        cliente_row = tk.Frame(left_inner, bg="white")
        cliente_row.pack(fill="x", padx=15, pady=(0, 5))
        
        tk.Label(cliente_row, text="Cliente:", font=("Segoe UI", 10), bg="white", fg="#374151").pack(side="left")
        
        self.pos_cliente = ttk.Combobox(cliente_row, state="readonly", font=("Segoe UI", 11))
        self.pos_cliente.pack(side="left", fill="x", expand=True, ipady=4, padx=(10, 5))
        self.cargar_clientes_pos()
        
        tk.Button(cliente_row, text="+", font=("Segoe UI", 10, "bold"), bg="#3b82f6", fg="white",
                 relief="flat", cursor="hand2", command=self.nuevo_cliente_desde_pos, width=3).pack(side="left")
        tk.Button(cliente_row, text="R", font=("Segoe UI", 10, "bold"), bg="#10b981", fg="white",
                 relief="flat", cursor="hand2", command=self.refrescar_clientes, width=3).pack(side="left", padx=3)
        
        self.pos_cliente_info = tk.Label(left_inner, text="", font=("Segoe UI", 9), 
                bg="white", fg="#64748b", anchor="w")
        self.pos_cliente_info.pack(fill="x", padx=15, pady=(0, 10))
        self.pos_cliente.bind("<<ComboboxSelected>>", lambda e: self.on_cliente_seleccionado())
        
        tk.Label(left_inner, text="METODO DE PAGO", font=("Segoe UI", 12, "bold"),
                bg="#fef3c7", fg="#92400e", anchor="w").pack(fill="x", padx=15, pady=(5, 0))
        
        pago_row = tk.Frame(left_inner, bg="white")
        pago_row.pack(fill="x", padx=15, pady=10)
        
        tk.Label(pago_row, text="Pago:", font=("Segoe UI", 10), bg="white", fg="#374151").pack(side="left")
        
        self.pos_pago = ttk.Combobox(pago_row, state="readonly", font=("Segoe UI", 11),
                                    values=["Efectivo", "Tarjeta Debito", "Tarjeta Credito", "Transferencia", "Yape/Plin", "Mixto"])
        self.pos_pago.current(0)
        self.pos_pago.pack(side="left", fill="x", expand=True, ipady=5, padx=(10, 5))
        
        tk.Label(pago_row, text="Ref:", font=("Segoe UI", 10), bg="white", fg="#374151").pack(side="left")
        
        self.pos_referencia = ttk.Entry(pago_row, font=("Segoe UI", 11), width=12)
        self.pos_referencia.pack(side="left", ipady=4, padx=(5, 0))
        
        tk.Label(left_inner, text="CARRITO DE COMPRAS", font=("Segoe UI", 12, "bold"),
                bg="#dbeafe", fg="#1e40af", anchor="w").pack(fill="x", padx=15, pady=(5, 0))
        
        self.cart_count_label = tk.Label(left_inner, text="0 items", font=("Segoe UI", 10),
                bg="#dbeafe", fg="#3b82f6", anchor="e")
        self.cart_count_label.pack(fill="x", padx=15)
        
        tree_frame = tk.Frame(left_inner, bg="white")
        tree_frame.pack(fill="both", expand=True, padx=15, pady=10)
        
        columns_cart = ("Producto", "Precio", "Cant", "Impuesto", "Subtotal")
        self.cart_tree = ttk.Treeview(tree_frame, columns=columns_cart, show="headings", height=8)
        
        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=28)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#f1f5f9", foreground="#1e293b")
        style.map("Treeview", background=[("selected", "#3b82f6")], foreground=[("selected", "white")])
        
        self.cart_tree.heading("Producto", text="PRODUCTO")
        self.cart_tree.heading("Precio", text="PRECIO")
        self.cart_tree.heading("Cant", text="CANT")
        self.cart_tree.heading("Impuesto", text="IMP")
        self.cart_tree.heading("Subtotal", text="SUBTOTAL")
        
        self.cart_tree.column("Producto", width=200)
        self.cart_tree.column("Precio", width=80, anchor="e")
        self.cart_tree.column("Cant", width=50, anchor="center")
        self.cart_tree.column("Impuesto", width=50, anchor="center")
        self.cart_tree.column("Subtotal", width=90, anchor="e")
        
        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.cart_tree.yview)
        self.cart_tree.configure(yscrollcommand=scroll.set)
        self.cart_tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        
        self.cart_tree.bind("<Double-1>", lambda e: self.editar_cantidad_carrito())
        
        btn_actions = tk.Frame(left_inner, bg="white")
        btn_actions.pack(fill="x", padx=15, pady=(0, 10))
        
        tk.Button(btn_actions, text="+", bg="#16a34a", fg="white", command=self.aumentar_cantidad, 
                 font=("Segoe UI", 14, "bold"), relief="flat", cursor="hand2", width=4, height=1).pack(side="left", padx=4)
        tk.Button(btn_actions, text="-", bg="#ea580c", fg="white", command=self.disminuir_cantidad,
                 font=("Segoe UI", 14, "bold"), relief="flat", cursor="hand2", width=4, height=1).pack(side="left", padx=4)
        tk.Button(btn_actions, text="X", bg="#dc2626", fg="white", command=self.eliminar_carrito,
                 font=("Segoe UI", 14, "bold"), relief="flat", cursor="hand2", width=4, height=1).pack(side="left", padx=4)
        tk.Button(btn_actions, text="Limpiar", bg="#6b7280", fg="white", command=self.vaciar_carrito, 
                 font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2", padx=15, pady=5).pack(side="right")
        
        totals = tk.Frame(left_inner, bg="#f8fafc", relief="solid", bd=1)
        totals.pack(fill="x", padx=15, pady=(0, 15))
        
        row1 = tk.Frame(totals, bg="#f8fafc")
        row1.pack(fill="x", padx=15, pady=(10, 5))
        tk.Label(row1, text="Subtotal:", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#475569").pack(side="left")
        self.subtotal_label = tk.Label(row1, text=curr["simbolo_local"] + " 0.00", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#1e293b")
        self.subtotal_label.pack(side="right")
        
        row2 = tk.Frame(totals, bg="#f8fafc")
        row2.pack(fill="x", padx=15, pady=(0, 5))
        tk.Label(row2, text="Impuesto (15% ISV):", font=("Segoe UI", 12, "bold"), bg="#f8fafc", fg="#475569").pack(side="left")
        self.impuesto_label = tk.Label(row2, text=curr["simbolo_local"] + " 0.00", font=("Segoe UI", 14, "bold"), bg="#f8fafc", fg="#1e293b")
        self.impuesto_label.pack(side="right")
        
        row3 = tk.Frame(totals, bg="#16a34a")
        row3.pack(fill="x", pady=(10, 0))
        
        tk.Label(row3, text="TOTAL:", font=("Segoe UI", 14, "bold"), bg="#16a34a", fg="white").pack(side="left", padx=15, pady=10)
        self.total_label = tk.Label(row3, text=curr["simbolo_local"] + " 0.00", font=("Segoe UI", 18, "bold"), bg="#16a34a", fg="white")
        self.total_label.pack(side="right", padx=15, pady=10)
        
        header_prod = tk.Frame(right_scroll_frame, bg="#1e293b")
        header_prod.pack(fill="x", padx=3, pady=(3, 0))
        
        tk.Label(header_prod, text="AGREGAR PRODUCTOS", font=("Segoe UI", 12, "bold"),
                bg="#1e293b", fg="white", anchor="w").pack(fill="x", padx=15, pady=10)
        
        prod_body = tk.Frame(right_scroll_frame, bg="white")
        prod_body.pack(fill="both", expand=True, padx=3)
        
        tk.Button(prod_body, text="VER CATALOGO", font=("Segoe UI", 11, "bold"),
                 bg="#2563eb", fg="white", relief="flat", cursor="hand2",
                 command=self.abrir_modal_productos, pady=10).pack(fill="x", padx=15, pady=(10, 10))
        
        tk.Label(prod_body, text="BUSQUEDA RAPIDA", font=("Segoe UI", 11, "bold"),
                bg="white", fg="#1e293b", anchor="w").pack(anchor="w", padx=15, pady=(0, 5))
        
        self.pos_entry = ttk.Entry(prod_body, font=("Segoe UI", 12))
        self.pos_entry.pack(fill="x", ipady=6, padx=15, pady=5)
        self.pos_entry.bind("<KeyRelease>", lambda e: self.buscar_productos_pos())
        self.pos_entry.bind("<Return>", lambda e: self.agregar_producto_rapido())
        self.pos_entry.focus()
        
        tree_frame = tk.Frame(prod_body, bg="white")
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(10, 5))
        
        columns_pos = ("Codigo", "Producto", "Precio", "Imp", "Stock")
        self.pos_tree = ttk.Treeview(tree_frame, columns=columns_pos, show="headings", height=6)
        
        self.pos_tree.heading("Codigo", text="CODIGO")
        self.pos_tree.heading("Producto", text="PRODUCTO")
        self.pos_tree.heading("Precio", text="PRECIO")
        self.pos_tree.heading("Imp", text="IMP")
        self.pos_tree.heading("Stock", text="STOCK")
        
        self.pos_tree.column("Codigo", width=80, anchor="center")
        self.pos_tree.column("Producto", width=140)
        self.pos_tree.column("Precio", width=70, anchor="e")
        self.pos_tree.column("Imp", width=40, anchor="center")
        self.pos_tree.column("Stock", width=50, anchor="center")
        
        pos_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.pos_tree.yview)
        self.pos_tree.configure(yscrollcommand=pos_scroll.set)
        self.pos_tree.pack(side="left", fill="both", expand=True)
        pos_scroll.pack(side="right", fill="y")
        
        self.pos_tree.tag_configure("bajo", background="#ffebee")
        self.pos_tree.bind("<Double-1>", lambda e: self.agregar_al_carrito())
        
        self.cargar_productos_pos()
        
        tk.Label(prod_body, text="ENTER o doble clic para agregar", font=("Segoe UI", 9),
                bg="white", fg="#9ca3af", anchor="w").pack(anchor="w", padx=15, pady=(5, 0))
        
        self.pos_stock_info = tk.Label(prod_body, text="", font=("Segoe UI", 10),
                bg="white", fg="#374151", anchor="w")
        self.pos_stock_info.pack(anchor="w", padx=15, pady=(5, 10))
        
    # ==================== METODOS DE CARGA DE DATOS ====================
    
    def cargar_proveedores(self):
        for item in self.prov_tree.get_children():
            self.prov_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        texto = self.prov_entry.get().strip()
        if texto:
            cursor.execute("SELECT * FROM proveedores WHERE nombre LIKE ? OR ruc_nit LIKE ? ORDER BY nombre", (f"%{texto}%", f"%{texto}%"))
        else:
            cursor.execute("SELECT * FROM proveedores ORDER BY nombre")
        for row in cursor.fetchall():
            self.prov_tree.insert("", "end", values=(row["id_proveedor"], row["nombre"], row["ruc_nit"], row["telefono"] or "-", row["email"] or "-", row["estado"]))
        conn.close()
        
    def cargar_contratos(self):
        from datetime import datetime
        for item in self.contrato_tree.get_children():
            self.contrato_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT c.*, p.nombre as proveedor FROM contratos c JOIN proveedores p ON c.id_proveedor = p.id_proveedor ORDER BY c.fecha_inicio DESC")
        
        hoy = datetime.now().date()
        dias_por_vencer = 30
        
        for row in cursor.fetchall():
            fecha_fin = datetime.strptime(row["fecha_fin"], "%Y-%m-%d").date()
            dias_restantes = (fecha_fin - hoy).days
            
            if row["estado"] == "vencido" or dias_restantes < 0:
                tag = "estado_vencido"
                alerta = "VENCIDO"
            elif dias_restantes <= dias_por_vencer:
                tag = "estado_por_vencer"
                alerta = f"{-dias_restantes}d" if dias_restantes <= 0 else f"{dias_restantes}d"
            else:
                tag = "estado_ok"
                alerta = "OK"
            
            self.contrato_tree.insert("", "end", tags=(tag,), values=(
                row["id_contrato"], 
                row["proveedor"], 
                row["fecha_inicio"], 
                row["fecha_fin"], 
                f"{row['porcentaje_comision']}%", 
                row["estado"],
                alerta
            ))
        conn.close()
        
    def cargar_espacios(self):
        for item in self.espacio_tree.get_children():
            self.espacio_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM espacios_exhibicion ORDER BY codigo")
        for row in cursor.fetchall():
            self.espacio_tree.insert("", "end", values=(row["id_espacio"], row["codigo"], row["descripcion"] or "-", row["ubicacion"] or "-", row["capacidad_max"], row["estado"]))
        conn.close()
        
    def cargar_productos(self):
        for item in self.prod_tree.get_children():
            self.prod_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        texto = self.prod_entry.get().strip()
        if texto:
            cursor.execute("""
                SELECT p.*, pr.nombre as proveedor, 
                       COALESCE((SELECT SUM(cant_actual) FROM inventarios WHERE id_producto = p.id_producto), 0) as stock_total
                FROM productos p 
                JOIN proveedores pr ON p.id_proveedor = pr.id_proveedor 
                WHERE p.nombre LIKE ? OR p.codigo_barras LIKE ? 
                ORDER BY p.nombre
            """, (f"%{texto}%", f"%{texto}%"))
        else:
            cursor.execute("""
                SELECT p.*, pr.nombre as proveedor,
                       COALESCE((SELECT SUM(cant_actual) FROM inventarios WHERE id_producto = p.id_producto), 0) as stock_total
                FROM productos p 
                JOIN proveedores pr ON p.id_proveedor = pr.id_proveedor 
                ORDER BY p.nombre
            """)
        
        curr = self.get_currency_info()
        for row in cursor.fetchall():
            row_dict = dict(row)
            precio_local = f"{curr['simbolo_local']} {row_dict['precio_venta']:,.2f}"
            tipo_imp = row_dict.get("tipo_impuesto", "gravado")
            tipo_label = {"gravado": "Gravado", "exento": "Exento", "zero": "Cero"}.get(tipo_imp, "Gravado")
            stock = row_dict["stock_total"]
            
            if stock == 0:
                self.prod_tree.insert("", "end", tags=("stock_bajo",), values=(row_dict["id_producto"], row_dict["codigo_barras"], row_dict["nombre"], row_dict["proveedor"], precio_local, tipo_label, stock, row_dict["estado"]))
            else:
                self.prod_tree.insert("", "end", values=(row_dict["id_producto"], row_dict["codigo_barras"], row_dict["nombre"], row_dict["proveedor"], precio_local, tipo_label, stock, row_dict["estado"]))
        conn.close()
        
    def cargar_inventario(self):
        for item in self.inv_tree.get_children():
            self.inv_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.*, p.nombre as producto, e.codigo as espacio,
                   CASE WHEN i.cant_actual <= i.cant_minima THEN 'BAJO' ELSE 'OK' END as estado_nivel
            FROM inventarios i JOIN productos p ON i.id_producto = p.id_producto
            JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio ORDER BY i.cant_actual ASC
        """)
        for row in cursor.fetchall():
            tag = "alerta" if row["estado_nivel"] == "BAJO" else "ok"
            self.inv_tree.insert("", "end", tags=(tag,), values=(row["id_inventario"], row["producto"], row["espacio"], row["cant_actual"], row["cant_minima"], row["cant_maxima"], row["estado_nivel"]))
        conn.close()
        
    def cargar_surtido(self):
        for item in self.surtido_tree.get_children():
            self.surtido_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT ps.*, p.nombre as producto, e.codigo as espacio FROM planif_surtido ps JOIN productos p ON ps.id_producto = p.id_producto JOIN espacios_exhibicion e ON ps.id_espacio = e.id_espacio ORDER BY ps.fecha_creacion DESC")
        for row in cursor.fetchall():
            self.surtido_tree.insert("", "end", values=(row["id_planificacion"], row["producto"], row["espacio"], row["cant_solicitada"], row["fecha_planificada"], row["estado"]))
        conn.close()
        
    def cargar_clientes(self):
        for item in self.cliente_tree.get_children():
            self.cliente_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        texto = self.cliente_entry.get().strip()
        if texto:
            cursor.execute("SELECT * FROM clientes WHERE nombre LIKE ? OR apellido LIKE ? OR doc_identidad LIKE ? ORDER BY nombre", (f"%{texto}%", f"%{texto}%", f"%{texto}%"))
        else:
            cursor.execute("SELECT * FROM clientes ORDER BY nombre")
        for row in cursor.fetchall():
            self.cliente_tree.insert("", "end", values=(row["id_cliente"], row["nombre"], row["apellido"] or "-", row["doc_identidad"] or "-", row["telefono"] or "-", row["email"] or "-"))
        conn.close()
        
    def cargar_liquidaciones(self):
        for item in self.liq_tree.get_children():
            self.liq_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT l.*, p.nombre as proveedor FROM liquidaciones l JOIN proveedores p ON l.id_proveedor = p.id_proveedor ORDER BY l.fecha_liquidacion DESC")
        for row in cursor.fetchall():
            self.liq_tree.insert("", "end", values=(row["id_liquidacion"], row["proveedor"], f"{row['periodo_inicio']} - {row['periodo_fin']}", f"S/. {row['total_ventas']:.2f}", f"S/. {row['monto_comision']:.2f}", f"S/. {row['monto_pagar']:.2f}", row["estado"]))
        conn.close()
        
    def cargar_productos_pos(self):
        for item in self.pos_tree.get_children():
            self.pos_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, i.cant_actual, i.id_inventario, i.id_producto, i.id_espacio,
                   COALESCE(SUM(i.cant_actual), 0) as stock_total
            FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
            WHERE p.estado = 'activo' AND i.cant_actual > 0
            GROUP BY p.id_producto ORDER BY p.nombre
        """)
        self.pos_productos = []
        curr = self.get_currency_info()
        for row in cursor.fetchall():
            row_dict = dict(row)
            precio_local = f"{curr['simbolo_local']} {row_dict['precio_venta']:,.2f}"
            tipo_imp = row_dict.get("tipo_impuesto", "gravado")
            tipo_label = {"gravado": "G", "exento": "E", "zero": "Z"}.get(tipo_imp, "G")
            stock = row_dict["cant_actual"]
            
            if stock <= 5:
                stock_text = f"{stock} !"
                self.pos_tree.insert("", "end", tags=("bajo",), values=(row_dict["codigo_barras"], row_dict["nombre"], precio_local, tipo_label, stock_text))
            else:
                self.pos_tree.insert("", "end", values=(row_dict["codigo_barras"], row_dict["nombre"], precio_local, tipo_label, stock))
            self.pos_productos.append(row_dict)
        
        self.pos_tree.tag_configure("bajo", background="#ffebee")
        conn.close()
        
    def buscar_productos_pos(self):
        texto = self.pos_entry.get().strip()
        for item in self.pos_tree.get_children():
            self.pos_tree.delete(item)
        conn = get_connection()
        cursor = conn.cursor()
        curr = self.get_currency_info()
        if texto:
            cursor.execute("""
                SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, SUM(i.cant_actual) as cant_actual, i.id_inventario, i.id_producto, i.id_espacio
                FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
                WHERE p.estado = 'activo' AND i.cant_actual > 0 AND (p.nombre LIKE ? OR p.codigo_barras LIKE ?)
                GROUP BY p.id_producto ORDER BY p.nombre
            """, (f"%{texto}%", f"%{texto}%"))
        else:
            cursor.execute("""
                SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, SUM(i.cant_actual) as cant_actual, i.id_inventario, i.id_producto, i.id_espacio
                FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
                WHERE p.estado = 'activo' AND i.cant_actual > 0
                GROUP BY p.id_producto ORDER BY p.nombre
            """)
        self.pos_productos = []
        for row in cursor.fetchall():
            row_dict = dict(row)
            precio_local = f"{curr['simbolo_local']} {row_dict['precio_venta']:,.2f}"
            tipo_imp = row_dict.get("tipo_impuesto", "gravado")
            tipo_label = {"gravado": "G", "exento": "E", "zero": "Z"}.get(tipo_imp, "G")
            stock = row_dict["cant_actual"]
            
            if stock <= 5:
                stock_text = f"{stock} !"
                self.pos_tree.insert("", "end", tags=("bajo",), values=(row_dict["codigo_barras"], row_dict["nombre"], precio_local, tipo_label, stock_text))
            else:
                self.pos_tree.insert("", "end", values=(row_dict["codigo_barras"], row_dict["nombre"], precio_local, tipo_label, stock))
            self.pos_productos.append(row_dict)
        
        self.pos_tree.tag_configure("bajo", background="#ffebee")
        conn.close()
        
    def cargar_clientes_pos(self):
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_cliente, nombre, apellido, doc_identidad FROM clientes ORDER BY nombre")
        clientes = cursor.fetchall()
        conn.close()
        valores = ["[SIN CLIENTE] Consumidor Final"]
        for c in clientes:
            nombre = f"{c['nombre']} {c['apellido'] or ''}".strip()
            doc = c['doc_identidad'] or 'S/D'
            valores.append(f"{nombre} | DOC: {doc}")
        self.pos_cliente["values"] = valores
        self.pos_cliente.current(0)
        
        self.pos_cliente.bind("<<ComboboxSelected>>", lambda e: self.on_cliente_seleccionado())
    
    def on_cliente_seleccionado(self):
        cliente = self.pos_cliente.get()
        if not cliente or cliente.startswith("[SIN CLIENTE]"):
            self.pos_cliente_info.config(text="Venta como Consumidor Final")
            return
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT c.*, COUNT(v.id_venta) as num_compras, COALESCE(SUM(v.total), 0) as total_compras
            FROM clientes c
            LEFT JOIN ventas v ON c.id_cliente = v.id_cliente
            WHERE c.nombre || ' ' || COALESCE(c.apellido, '') || ' | DOC: ' || COALESCE(c.doc_identidad, 'S/D') = ?
            GROUP BY c.id_cliente
        """, (cliente,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            info = f"Cliente: {cliente.split(' | ')[0]} | Doc: {result['doc_identidad'] or 'S/D'} | Compras: {result['num_compras']} | Total: L {result['total_compras']:,.2f}"
            self.pos_cliente_info.config(text=info)
    
    def refrescar_clientes(self):
        self.cargar_clientes_pos()
        self.pos_cliente_info.config(text="")
        
    def nuevo_cliente_desde_pos(self):
        form = tk.Toplevel(self.root)
        form.title("Nuevo Cliente")
        form.geometry("400x320")
        form.configure(bg="white")
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg="#3b82f6")
        header.pack(fill="x")
        tk.Label(header, text="Agregar Nuevo Cliente", font=("Segoe UI", 14, "bold"),
                bg="#3b82f6", fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        campos = [("Nombre:", "nombre"), ("Apellido:", "apellido"), ("Telefono:", "telefono"), ("Email:", "email")]
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="w", pady=8, padx=(0, 10))
            entry = ttk.Entry(card, width=30, font=("Segoe UI", 11))
            entry.grid(row=i, column=1, pady=8)
            entries[key] = entry
        
        def guardar():
            if not entries["nombre"].get().strip():
                messagebox.showwarning("Error", "El nombre es obligatorio")
                return
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO clientes (nombre, apellido, telefono, email) VALUES (?, ?, ?, ?)",
                             (entries["nombre"].get(), entries["apellido"].get(), entries["telefono"].get(), entries["email"].get()))
                conn.commit()
                form.destroy()
                self.cargar_clientes_pos()
                messagebox.showinfo("Exito", "Cliente agregado correctamente")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
        
        tk.Button(card, text="Guardar Cliente", font=("Segoe UI", 11, "bold"), bg="#3b82f6", fg="white",
                 relief="flat", cursor="hand2", padx=20, pady=10, command=guardar).grid(row=len(campos), column=0, columnspan=2, pady=20)
        
    # ==================== ACCIONES CRUD ====================
    
    def nuevo_proveedor(self):
        self.mostrar_formulario("Nuevo Proveedor", self.S.COLORS["success"], self.guardar_proveedor)
        
    def editar_proveedor(self):
        selection = self.prov_tree.selection()
        if not selection:
            messagebox.showwarning("Seleccionar", "Seleccione un proveedor")
            return
        item = self.prov_tree.item(selection[0])
        self.mostrar_formulario("Editar Proveedor", self.S.COLORS["accent"], self.guardar_proveedor, item["values"][0])
        
    def mostrar_formulario(self, titulo, color, comando, id_editar=None):
        form = tk.Toplevel(self.root)
        form.title(titulo)
        form.geometry("450x400")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=color)
        header.pack(fill="x")
        tk.Label(header, text=titulo, font=("Segoe UI", 14, "bold"), bg=color, fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        campos = [
            ("Nombre:", "nombre"),
            ("RUC/NIT:", "ruc_nit"),
            ("Telefono:", "telefono"),
            ("Email:", "email"),
        ]
        
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="w", pady=10)
            entry = ttk.Entry(card, width=35, font=("Segoe UI", 11))
            entry.grid(row=i, column=1, pady=10, padx=10)
            entries[key] = entry
            
        if id_editar:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM proveedores WHERE id_proveedor=?", (id_editar,))
            datos = cursor.fetchone()
            conn.close()
            if datos:
                entries["nombre"].insert(0, datos["nombre"])
                entries["ruc_nit"].insert(0, datos["ruc_nit"])
                entries["telefono"].insert(0, datos["telefono"] or "")
                entries["email"].insert(0, datos["email"] or "")
                
        def guardar():
            conn = get_connection()
            cursor = conn.cursor()
            try:
                if id_editar:
                    cursor.execute("UPDATE proveedores SET nombre=?, ruc_nit=?, telefono=?, email=? WHERE id_proveedor=?", 
                                 (entries["nombre"].get(), entries["ruc_nit"].get(), entries["telefono"].get(), entries["email"].get(), id_editar))
                else:
                    cursor.execute("INSERT INTO proveedores (nombre, ruc_nit, telefono, email) VALUES (?, ?, ?, ?)",
                                 (entries["nombre"].get(), entries["ruc_nit"].get(), entries["telefono"].get(), entries["email"].get()))
                conn.commit()
                form.destroy()
                self.cargar_proveedores()
                messagebox.showinfo("Exito", "Guardado correctamente")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=color, fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=5, column=0, columnspan=2, pady=25)
        
    def guardar_proveedor(self):
        pass
        
    def nuevo_contrato(self):
        form = tk.Toplevel(self.root)
        form.title("Nuevo Contrato")
        form.geometry("400x320")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        tk.Label(header, text="Nuevo Contrato", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["primary"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT nombre FROM proveedores WHERE estado='activo'")
        proveedores = [p["nombre"] for p in cursor.fetchall()]
        conn.close()
        
        tk.Label(card, text="Proveedor:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", pady=10)
        prov_var = tk.StringVar()
        ttk.Combobox(card, textvariable=prov_var, values=proveedores, width=33, state="readonly").grid(row=0, column=1, pady=10, padx=10)
        
        tk.Label(card, text="Fecha Inicio:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=1, column=0, sticky="w", pady=10)
        entry_inicio = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        entry_inicio.grid(row=1, column=1, pady=10, padx=10)
        entry_inicio.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        tk.Label(card, text="Fecha Fin:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", pady=10)
        entry_fin = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        entry_fin.grid(row=2, column=1, pady=10, padx=10)
        entry_fin.insert(0, (datetime.now().replace(day=1) + timedelta(days=365)).strftime("%Y-%m-%d"))
        
        tk.Label(card, text="Comision (%):", font=("Segoe UI", 10, "bold"), bg="white").grid(row=3, column=0, sticky="w", pady=10)
        entry_comision = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        entry_comision.grid(row=3, column=1, pady=10, padx=10)
        entry_comision.insert(0, "15.0")
        
        def guardar():
            if not prov_var.get():
                messagebox.showwarning("Error", "Seleccione un proveedor")
                return
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id_proveedor FROM proveedores WHERE nombre=?", (prov_var.get(),))
            id_prov = cursor.fetchone()["id_proveedor"]
            try:
                cursor.execute("INSERT INTO contratos (id_proveedor, fecha_inicio, fecha_fin, porcentaje_comision, condiciones) VALUES (?, ?, ?, ?, 'Contrato estandar')",
                             (id_prov, entry_inicio.get(), entry_fin.get(), float(entry_comision.get())))
                conn.commit()
                form.destroy()
                self.cargar_contratos()
                messagebox.showinfo("Exito", "Contrato creado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=4, column=0, columnspan=2, pady=25)
        
    def nuevo_espacio(self):
        form = tk.Toplevel(self.root)
        form.title("Nuevo Espacio")
        form.geometry("400x280")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        tk.Label(header, text="Nuevo Espacio", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["primary"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        campos = [("Codigo:", "codigo"), ("Descripcion:", "descripcion"), ("Ubicacion:", "ubicacion")]
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="w", pady=10)
            entry = ttk.Entry(card, width=30, font=("Segoe UI", 11))
            entry.grid(row=i, column=1, pady=10, padx=10)
            entries[key] = entry
            
        tk.Label(card, text="Capacidad:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=3, column=0, sticky="w", pady=10)
        entry_cap = ttk.Entry(card, width=30, font=("Segoe UI", 11))
        entry_cap.grid(row=3, column=1, pady=10, padx=10)
        entry_cap.insert(0, "100")
        
        def guardar():
            if not entries["codigo"].get().strip():
                messagebox.showwarning("Error", "El codigo es obligatorio")
                return
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO espacios_exhibicion (codigo, descripcion, ubicacion, capacidad_max) VALUES (?, ?, ?, ?)",
                             (entries["codigo"].get(), entries["descripcion"].get(), entries["ubicacion"].get(), int(entry_cap.get())))
                conn.commit()
                form.destroy()
                self.cargar_espacios()
                messagebox.showinfo("Exito", "Espacio creado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=4, column=0, columnspan=2, pady=15)
        
    def nuevo_producto(self):
        form = tk.Toplevel(self.root)
        form.title("Nuevo Producto")
        form.geometry("450x450")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["purple"])
        header.pack(fill="x")
        tk.Label(header, text="Nuevo Producto", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["purple"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_proveedor, nombre FROM proveedores WHERE estado='activo'")
        proveedores = list(cursor.fetchall())
        conn.close()
        
        campos = {}
        
        tk.Label(card, text="Proveedor:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", pady=8)
        campos["proveedor"] = tk.StringVar()
        ttk.Combobox(card, textvariable=campos["proveedor"], values=[p["nombre"] for p in proveedores], width=33, state="readonly").grid(row=0, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Nombre:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=1, column=0, sticky="w", pady=8)
        campos["nombre"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["nombre"].grid(row=1, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Codigo:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", pady=8)
        campos["codigo"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["codigo"].grid(row=2, column=1, pady=8, padx=10)
        campos["codigo"].insert(0, f"750{random.randint(10000000, 99999999)}")
        
        tk.Label(card, text="Precio Venta:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=3, column=0, sticky="w", pady=8)
        campos["precio_venta"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["precio_venta"].grid(row=3, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Precio Costo:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=4, column=0, sticky="w", pady=8)
        campos["precio_costo"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["precio_costo"].grid(row=4, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Tipo Impuesto:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=5, column=0, sticky="w", pady=8)
        campos["tipo_impuesto"] = tk.StringVar(value="gravado")
        tipo_frame = tk.Frame(card, bg="white")
        tipo_frame.grid(row=5, column=1, pady=8, padx=10, sticky="w")
        tk.Radiobutton(tipo_frame, text="Gravado (15%)", variable=campos["tipo_impuesto"], value="gravado", bg="white", font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
        tk.Radiobutton(tipo_frame, text="Exento", variable=campos["tipo_impuesto"], value="exento", bg="white", font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
        tk.Radiobutton(tipo_frame, text="Cero", variable=campos["tipo_impuesto"], value="zero", bg="white", font=("Segoe UI", 9)).pack(side="left")
        
        def guardar():
            if not campos["proveedor"].get() or not campos["nombre"].get().strip():
                messagebox.showwarning("Error", "Complete los campos obligatorios")
                return
            id_prov = next((p["id_proveedor"] for p in proveedores if p["nombre"] == campos["proveedor"].get()), None)
            precio_venta = float(campos["precio_venta"].get()) if campos["precio_venta"].get() else 0
            precio_costo = float(campos["precio_costo"].get()) if campos["precio_costo"].get() else None
            tipo_imp = campos["tipo_impuesto"].get()
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO productos (id_proveedor, nombre, codigo_barras, precio_venta, precio_costo, tipo_impuesto, estado) VALUES (?, ?, ?, ?, ?, ?, 'activo')",
                             (id_prov, campos["nombre"].get().strip(), campos["codigo"].get(), precio_venta, precio_costo, tipo_imp))
                conn.commit()
                form.destroy()
                self.cargar_productos()
                messagebox.showinfo("Exito", "Producto guardado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=15, command=guardar).grid(row=6, column=0, columnspan=2, pady=20)
        
    def editar_producto(self):
        if not hasattr(self, 'prod_tree') or self.prod_tree is None:
            messagebox.showwarning("Error", "Primero vaya al modulo de Productos")
            return
        selection = self.prod_tree.selection()
        if not selection:
            messagebox.showwarning("Sin seleccion", "Seleccione un producto")
            return
        
        item = self.prod_tree.item(selection[0])
        values = item["values"]
        id_producto = values[0]
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM productos WHERE id_producto = ?", (id_producto,))
        producto = cursor.fetchone()
        
        cursor.execute("SELECT id_proveedor, nombre FROM proveedores WHERE estado='activo'")
        proveedores = list(cursor.fetchall())
        conn.close()
        
        if not producto:
            messagebox.showerror("Error", "Producto no encontrado")
            return
        
        form = tk.Toplevel(self.root)
        form.title("Editar Producto")
        form.geometry("450x450")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["purple"])
        header.pack(fill="x")
        tk.Label(header, text="Editar Producto", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["purple"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        campos = {}
        
        tk.Label(card, text="Proveedor:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", pady=8)
        campos["proveedor"] = tk.StringVar()
        prov_combo = ttk.Combobox(card, textvariable=campos["proveedor"], values=[p["nombre"] for p in proveedores], width=33, state="readonly")
        prov_combo.grid(row=0, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Nombre:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=1, column=0, sticky="w", pady=8)
        campos["nombre"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["nombre"].grid(row=1, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Codigo:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", pady=8)
        campos["codigo"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["codigo"].grid(row=2, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Precio Venta:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=3, column=0, sticky="w", pady=8)
        campos["precio_venta"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["precio_venta"].grid(row=3, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Precio Costo:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=4, column=0, sticky="w", pady=8)
        campos["precio_costo"] = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        campos["precio_costo"].grid(row=4, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Estado:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=5, column=0, sticky="w", pady=8)
        campos["estado"] = tk.StringVar()
        estado_combo = ttk.Combobox(card, textvariable=campos["estado"], values=["activo", "inactivo", "agotado"], width=33, state="readonly")
        estado_combo.grid(row=5, column=1, pady=8, padx=10)
        
        prov_nombre = next((p["nombre"] for p in proveedores if p["id_proveedor"] == producto["id_proveedor"]), "")
        campos["proveedor"].set(prov_nombre)
        campos["nombre"].insert(0, producto["nombre"])
        campos["codigo"].insert(0, producto["codigo_barras"] or "")
        campos["precio_venta"].insert(0, str(producto["precio_venta"]))
        if producto["precio_costo"]:
            campos["precio_costo"].insert(0, str(producto["precio_costo"]))
        campos["estado"].set(producto["estado"])
        
        def guardar():
            if not campos["proveedor"].get() or not campos["nombre"].get().strip():
                messagebox.showwarning("Error", "Complete los campos obligatorios")
                return
            id_prov = next((p["id_proveedor"] for p in proveedores if p["nombre"] == campos["proveedor"].get()), None)
            precio_venta = float(campos["precio_venta"].get()) if campos["precio_venta"].get() else 0
            precio_costo = float(campos["precio_costo"].get()) if campos["precio_costo"].get() else None
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    UPDATE productos SET id_proveedor=?, nombre=?, codigo_barras=?, precio_venta=?, precio_costo=?, estado=? WHERE id_producto=?
                """, (id_prov, campos["nombre"].get().strip(), campos["codigo"].get(), precio_venta, precio_costo, campos["estado"].get(), id_producto))
                conn.commit()
                form.destroy()
                self.cargar_productos()
                messagebox.showinfo("Exito", "Producto actualizado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Actualizar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=15, command=guardar).grid(row=6, column=0, columnspan=2, pady=20)
        
    def exportar_productos_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.id_producto, p.codigo_barras, p.nombre, pr.nombre as proveedor, 
                       p.precio_venta, p.precio_costo, p.estado, p.fecha_registro
                FROM productos p
                JOIN proveedores pr ON p.id_proveedor = pr.id_proveedor
                ORDER BY p.nombre
            """)
            productos = cursor.fetchall()
            conn.close()
            
            curr = self.get_currency_info()
            simbolo = curr['simbolo_local']
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos"
            
            header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ["ID", "Codigo", "Nombre", "Proveedor", "P. Venta", "P. Costo", "Estado", "Fecha Registro"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row_idx, prod in enumerate(productos, 2):
                ws.cell(row=row_idx, column=1, value=prod["id_producto"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=prod["codigo_barras"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=prod["nombre"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=prod["proveedor"]).border = thin_border
                
                cell_venta = ws.cell(row=row_idx, column=5, value=prod["precio_venta"])
                cell_venta.number_format = f'"{simbolo} "#,##0.00'
                cell_venta.border = thin_border
                
                cell_costo = ws.cell(row=row_idx, column=6, value=prod["precio_costo"] if prod["precio_costo"] else 0)
                cell_costo.number_format = f'"{simbolo} "#,##0.00'
                cell_costo.border = thin_border
                
                ws.cell(row=row_idx, column=7, value=prod["estado"]).border = thin_border
                ws.cell(row=row_idx, column=8, value=prod["fecha_registro"]).border = thin_border
            
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 25
            
            ws.row_dimensions[1].height = 25
            
            filename = f"productos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
            wb.save(filepath)
            
            messagebox.showinfo("Exportado", f"Archivo guardado:\n{filename}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar:\n{str(e)}")
        
    def asignar_inventario(self):
        form = tk.Toplevel(self.root)
        form.title("Asignar Producto")
        form.geometry("450x420")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["success"])
        header.pack(fill="x")
        tk.Label(header, text="Asignar Producto a Espacio", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["success"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_producto, nombre FROM productos WHERE estado='activo'")
        productos = [f"{p['id_producto']} - {p['nombre']}" for p in cursor.fetchall()]
        
        cursor.execute("SELECT id_espacio, codigo, descripcion, ubicacion, capacidad_max FROM espacios_exhibicion ORDER BY codigo")
        espacios_data = list(cursor.fetchall())
        espacios = [f"{e['id_espacio']} - {e['codigo']}" for e in espacios_data]
        conn.close()
        
        tk.Label(card, text="Producto:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", pady=8)
        prod_var = tk.StringVar()
        prod_combo = ttk.Combobox(card, textvariable=prod_var, values=productos, width=35, state="readonly")
        prod_combo.grid(row=0, column=1, pady=8, padx=10)
        
        tk.Label(card, text="Espacio:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=1, column=0, sticky="w", pady=8)
        esp_var = tk.StringVar()
        esp_combo = ttk.Combobox(card, textvariable=esp_var, values=espacios, width=35, state="readonly")
        esp_combo.grid(row=1, column=1, pady=8, padx=10)
        
        info_frame = tk.Frame(card, bg="#f5f5f5", padx=10, pady=10)
        info_frame.grid(row=2, column=0, columnspan=2, pady=15, sticky="ew")
        
        capacidad_label = tk.Label(info_frame, text="Capacidad maxima: -", font=("Segoe UI", 10), bg="#f5f5f5", fg="#666")
        capacidad_label.pack(anchor="w")
        
        stock_label = tk.Label(info_frame, text="Stock actual: -", font=("Segoe UI", 10), bg="#f5f5f5", fg="#666")
        stock_label.pack(anchor="w")
        
        disponible_label = tk.Label(info_frame, text="Disponible: -", font=("Segoe UI", 11, "bold"), bg="#f5f5f5", fg="#27ae60")
        disponible_label.pack(anchor="w")
        
        def actualizar_info_espacio(*args):
            if not esp_var.get():
                capacidad_label.config(text="Capacidad maxima: -")
                stock_label.config(text="Stock actual: -")
                disponible_label.config(text="Disponible: -")
                return
            
            esp_id = int(esp_var.get().split(" - ")[0])
            
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT capacidad_max FROM espacios_exhibicion WHERE id_espacio = ?", (esp_id,))
            capacidad = cursor.fetchone()["capacidad_max"]
            
            cursor.execute("SELECT COALESCE(SUM(cant_actual), 0) as total FROM inventarios WHERE id_espacio = ?", (esp_id,))
            stock_actual = cursor.fetchone()["total"]
            
            disponible = max(0, capacidad - stock_actual)
            
            capacidad_label.config(text=f"Capacidad maxima: {capacidad} unidades")
            stock_label.config(text=f"Stock actual: {stock_actual} unidades")
            disponible_label.config(text=f"Disponible: {disponible} unidades")
            
            if disponible == 0:
                disponible_label.config(fg="#e74c3c")
            elif disponible < 10:
                disponible_label.config(fg="#f39c12")
            else:
                disponible_label.config(fg="#27ae60")
            
            conn.close()
        
        esp_var.trace("w", actualizar_info_espacio)
        
        tk.Label(card, text="Cantidad a asignar:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=3, column=0, sticky="w", pady=8)
        entry_cant = ttk.Entry(card, width=37, font=("Segoe UI", 11))
        entry_cant.grid(row=3, column=1, pady=8, padx=10)
        entry_cant.insert(0, "1")
        
        def guardar():
            if not prod_var.get() or not esp_var.get():
                messagebox.showwarning("Error", "Complete todos los campos")
                return
            
            cantidad = int(entry_cant.get()) if entry_cant.get() else 0
            if cantidad <= 0:
                messagebox.showwarning("Error", "La cantidad debe ser mayor a 0")
                return
            
            esp_id = int(esp_var.get().split(" - ")[0])
            
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT capacidad_max FROM espacios_exhibicion WHERE id_espacio = ?", (esp_id,))
            capacidad = cursor.fetchone()["capacidad_max"]
            
            cursor.execute("SELECT COALESCE(SUM(cant_actual), 0) as total FROM inventarios WHERE id_espacio = ?", (esp_id,))
            stock_actual = cursor.fetchone()["total"]
            
            disponible = capacidad - stock_actual
            
            if cantidad > disponible:
                messagebox.showwarning("Capacidad excedida", f"El espacio tiene {disponible} lugares disponibles.\nNo puede asignar {cantidad} unidades.")
                conn.close()
                return
            
            id_prod = int(prod_var.get().split(" - ")[0])
            
            try:
                cursor.execute("INSERT INTO inventarios (id_producto, id_espacio, cant_actual, cant_minima, cant_maxima) VALUES (?, ?, ?, 5, 100)",
                             (id_prod, esp_id, cantidad))
                conn.commit()
                form.destroy()
                self.cargar_inventario()
                messagebox.showinfo("Exito", f"Producto asignado al espacio.\n{cantidad} unidades agregadas.")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=4, column=0, columnspan=2, pady=20)
        
    def ajustar_stock(self):
        selection = self.inv_tree.selection()
        if not selection:
            messagebox.showwarning("Seleccionar", "Seleccione un registro")
            return
        item = self.inv_tree.item(selection[0])
        values = item["values"]
        
        form = tk.Toplevel(self.root)
        form.title("Ajuste de Stock")
        form.geometry("400x280")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["warning"])
        header.pack(fill="x")
        tk.Label(header, text="Ajuste de Stock", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["warning"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        tk.Label(card, text=f"Producto: {values[1]}", font=("Segoe UI", 12, "bold"), bg="white").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))
        tk.Label(card, text=f"Stock Actual: {values[3]}", font=("Segoe UI", 10), bg="white").grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 15))
        
        tk.Label(card, text="Nuevo Stock:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", pady=10)
        entry_nuevo = ttk.Entry(card, width=25, font=("Segoe UI", 11))
        entry_nuevo.grid(row=2, column=1, pady=10, padx=10)
        entry_nuevo.insert(0, values[3])
        
        def guardar():
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("UPDATE inventarios SET cant_actual=? WHERE id_inventario=?", (int(entry_nuevo.get()), values[0]))
                conn.commit()
                form.destroy()
                self.cargar_inventario()
                messagebox.showinfo("Exito", "Stock ajustado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["warning"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=3, column=0, columnspan=2, pady=15)
    
    def reposicion_inventario(self):
        form = tk.Toplevel(self.root)
        form.title("Reposición de Inventario")
        form.geometry("900x600")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg="#1e3a5f")
        header.pack(fill="x")
        tk.Label(header, text="Reposición de Inventario", font=("Segoe UI", 16, "bold"), 
                bg="#1e3a5f", fg="white").pack(pady=15, padx=20)
        tk.Label(header, text="Productos con stock bajo que tienen proveedor asignado", 
                font=("Segoe UI", 10), bg="#1e3a5f", fg="#94a3b8").pack(pady=(0, 10), padx=20)
        
        toolbar = tk.Frame(form, bg="white")
        toolbar.pack(fill="x", padx=20, pady=15)
        
        tk.Label(toolbar, text="Buscar:", bg="white", font=("Segoe UI", 10)).pack(side="left", padx=(0, 10))
        
        search_entry = ttk.Entry(toolbar, width=40, font=("Segoe UI", 11))
        search_entry.pack(side="left", ipady=6, padx=(0, 20))
        
        info_label = tk.Label(toolbar, text="", bg="white", font=("Segoe UI", 10), fg="#64748b")
        info_label.pack(side="right")
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        columns = ("ID", "Producto", "Proveedor", "Espacio", "Stock", "Min", "Max", "Solicitar")
        tree = ttk.Treeview(card, columns=columns, show="headings", height=15)
        
        for col in columns:
            tree.heading(col, text=col)
            widths = {"ID": 50, "Producto": 220, "Proveedor": 150, "Espacio": 80, "Stock": 60, "Min": 50, "Max": 50, "Solicitar": 80}
            tree.column(col, width=widths.get(col, 100), anchor="center")
        
        tree.tag_configure("critico", background="#fef2f2")
        tree.tag_configure("bajo", background="#fef3c7")
        tree.tag_configure("normal", background="#f0fdf4")
        
        scrollbar = ttk.Scrollbar(card, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        cantidad_entries = {}
        
        def cargar_productos():
            for item in tree.get_children():
                tree.delete(item)
            
            conn = get_connection()
            cursor = conn.cursor()
            
            texto = search_entry.get().strip()
            
            query = """
                SELECT i.id_inventario, p.id_producto, p.nombre as producto, p.codigo_barras,
                       pr.nombre as proveedor, e.codigo as espacio, 
                       i.cant_actual, i.cant_minima, i.cant_maxima,
                       i.cant_maxima - i.cant_actual as cant_necesaria
                FROM inventarios i 
                JOIN productos p ON i.id_producto = p.id_producto
                JOIN proveedores pr ON p.id_proveedor = pr.id_proveedor
                JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio
                WHERE p.estado = 'activo' AND pr.estado = 'activo'
                ORDER BY i.cant_actual ASC
            """
            
            if texto:
                query = f"""
                    SELECT i.id_inventario, p.id_producto, p.nombre as producto, p.codigo_barras,
                           pr.nombre as proveedor, e.codigo as espacio, 
                           i.cant_actual, i.cant_minima, i.cant_maxima,
                           i.cant_maxima - i.cant_actual as cant_necesaria
                    FROM inventarios i 
                    JOIN productos p ON i.id_producto = p.id_producto
                    JOIN proveedores pr ON p.id_proveedor = pr.id_proveedor
                    JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio
                    WHERE p.estado = 'activo' AND pr.estado = 'activo'
                      AND (p.nombre LIKE ? OR pr.nombre LIKE ? OR p.codigo_barras LIKE ?)
                    ORDER BY i.cant_actual ASC
                """
                texto_buscar = f"%{texto}%"
                cursor.execute(query, (texto_buscar, texto_buscar, texto_buscar))
            else:
                cursor.execute(query)
            
            productos = cursor.fetchall()
            total_bajo = 0
            
            for prod in productos:
                if prod["cant_actual"] <= prod["cant_minima"]:
                    total_bajo += 1
                    stock_ratio = prod["cant_actual"] / prod["cant_minima"] if prod["cant_minima"] > 0 else 0
                    
                    if stock_ratio <= 0.5 or prod["cant_actual"] == 0:
                        tag = "critico"
                    elif stock_ratio <= 1:
                        tag = "bajo"
                    else:
                        tag = "normal"
                    
                    cant_default = prod["cant_maxima"] - prod["cant_actual"]
                    cantidad_entries[prod["id_inventario"]] = cant_default
                    
                    tree.insert("", "end", tags=(tag,), values=(
                        prod["id_inventario"],
                        f"{prod['producto'][:35]}",
                        f"{prod['proveedor'][:20]}",
                        prod["espacio"],
                        prod["cant_actual"],
                        prod["cant_minima"],
                        prod["cant_maxima"],
                        cant_default
                    ))
            
            info_label.config(text=f"Productos con stock bajo: {total_bajo}")
            conn.close()
        
        def actualizar_cantidad(event):
            selection = tree.selection()
            if not selection:
                return
            item = tree.item(selection[0])
            values = list(item["values"])
            inv_id = values[0]
            
            inv_win = tk.Toplevel(form)
            inv_win.title("Cantidad a Solicitar")
            inv_win.geometry("350x200")
            inv_win.configure(bg=self.S.COLORS["bg_main"])
            inv_win.transient(form)
            inv_win.grab_set()
            inv_win.center_on_parent()
            
            card_win = tk.Frame(inv_win, bg="white")
            card_win.pack(fill="both", expand=True, padx=20, pady=20)
            
            tk.Label(card_win, text=f"Producto: {values[1]}", font=("Segoe UI", 11, "bold"),
                    bg="white").pack(anchor="w")
            tk.Label(card_win, text=f"Stock actual: {values[4]} | Máximo: {values[6]}",
                    font=("Segoe UI", 10), bg="white", fg="#64748b").pack(anchor="w", pady=(5, 15))
            
            tk.Label(card_win, text="Cantidad a solicitar:", font=("Segoe UI", 10), bg="white").pack(anchor="w")
            entry_cant = ttk.Entry(card_win, font=("Segoe UI", 14), width=15)
            entry_cant.insert(0, str(values[7]))
            entry_cant.pack(pady=10)
            entry_cant.select_range(0, tk.END)
            entry_cant.focus()
            
            def guardar_cantidad():
                try:
                    nueva_cant = int(entry_cant.get())
                    if nueva_cant < 0:
                        messagebox.showwarning("Error", "La cantidad no puede ser negativa")
                        return
                    cantidad_entries[inv_id] = nueva_cant
                    values[7] = nueva_cant
                    tree.item(selection[0], values=values)
                    inv_win.destroy()
                except ValueError:
                    messagebox.showwarning("Error", "Ingrese un número válido")
            
            btn_frame = tk.Frame(card_win, bg="white")
            btn_frame.pack(pady=10)
            tk.Button(btn_frame, text="Guardar", font=("Segoe UI", 10, "bold"),
                     bg=self.S.COLORS["success"], fg="white", relief="flat",
                     padx=20, pady=8, command=guardar_cantidad).pack(side="left", padx=5)
            tk.Button(btn_frame, text="Cancelar", font=("Segoe UI", 10),
                     bg="#e2e8f0", fg="#475569", relief="flat",
                     padx=20, pady=8, command=inv_win.destroy).pack(side="left", padx=5)
            
            entry_cant.bind("<Return>", lambda e: guardar_cantidad())
        
        tree.bind("<Double-1>", actualizar_cantidad)
        search_entry.bind("<KeyRelease>", lambda e: cargar_productos())
        
        btn_frame = tk.Frame(form, bg="white")
        btn_frame.pack(fill="x", padx=20, pady=15)
        
        def generar_pedido():
            items_seleccionados = []
            
            for item in tree.get_children():
                values = tree.item(item)["values"]
                inv_id = values[0]
                if inv_id in cantidad_entries:
                    cant = cantidad_entries[inv_id]
                    if cant > 0:
                        items_seleccionados.append({
                            "id_inventario": inv_id,
                            "producto": values[1],
                            "proveedor": values[2],
                            "espacio": values[3],
                            "cantidad": cant
                        })
            
            if not items_seleccionados:
                messagebox.showwarning("Sin selección", "No hay productos seleccionados para reponer")
                return
            
            confirmar = messagebox.askyesno("Confirmar Pedido", 
                f"Se crearán {len(items_seleccionados)} solicitudes de reposición.\n\n¿Continuar?")
            
            if not confirmar:
                return
            
            conn = get_connection()
            cursor = conn.cursor()
            creados = 0
            
            try:
                for item in items_seleccionados:
                    cursor.execute("""
                        INSERT INTO planif_surtido (id_producto, id_espacio, fecha_planificada, cant_solicitada, estado)
                        SELECT p.id_producto, i.id_espacio, ?, ?, 'pendiente'
                        FROM productos p
                        JOIN inventarios i ON p.id_producto = i.id_producto
                        WHERE i.id_inventario = ?
                    """, (datetime.now().strftime("%Y-%m-%d"), item["cantidad"], item["id_inventario"]))
                    creados += 1
                
                conn.commit()
                messagebox.showinfo("Éxito", f"Se crearon {creados} solicitudes de reposición")
                cargar_productos()
            except Exception as e:
                messagebox.showerror("Error", str(e))
                conn.rollback()
            finally:
                conn.close()
        
        def ejecutar_reposicion():
            items_seleccionados = []
            
            for item in tree.get_children():
                values = tree.item(item)["values"]
                inv_id = values[0]
                if inv_id in cantidad_entries:
                    cant = cantidad_entries[inv_id]
                    if cant > 0:
                        items_seleccionados.append({
                            "id_inventario": inv_id,
                            "producto": values[1],
                            "proveedor": values[2],
                            "espacio": values[3],
                            "stock_actual": values[4],
                            "cantidad": cant
                        })
            
            if not items_seleccionados:
                messagebox.showwarning("Sin selección", "No hay productos seleccionados para reponer")
                return
            
            confirmar = messagebox.askyesno("Confirmar Reposición Directa", 
                f"Se agregarán {len(items_seleccionados)} productos al inventario.\n\n¿Continuar?")
            
            if not confirmar:
                return
            
            conn = get_connection()
            cursor = conn.cursor()
            actualizados = 0
            
            try:
                for item in items_seleccionados:
                    cursor.execute("""
                        UPDATE inventarios 
                        SET cant_actual = cant_actual + ?,
                            f_ultimo_surtido = ?
                        WHERE id_inventario = ?
                    """, (item["cantidad"], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), item["id_inventario"]))
                    
                    cursor.execute("""
                        INSERT INTO historial_inventario (id_inventario, tipo_movimiento, cantidad, observaciones, usuario)
                        VALUES (?, 'entrada', ?, 'Reposición directa', ?)
                    """, (item["id_inventario"], item["cantidad"], self.usuario_actual["id_usuario"]))
                    
                    actualizados += 1
                
                conn.commit()
                messagebox.showinfo("Éxito", f"Se actualizaron {actualizados} productos en el inventario")
                cargar_productos()
                self.cargar_inventario()
            except Exception as e:
                messagebox.showerror("Error", str(e))
                conn.rollback()
            finally:
                conn.close()
        
        tk.Button(btn_frame, text="Generar Solicitud de Reposición", font=("Segoe UI", 11, "bold"),
                 bg="#7c3aed", fg="white", relief="flat", cursor="hand2",
                 padx=20, pady=10, command=generar_pedido).pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="Reposición Directa (+ Stock)", font=("Segoe UI", 11, "bold"),
                 bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2",
                 padx=20, pady=10, command=ejecutar_reposicion).pack(side="left", padx=5)
        
        tk.Label(btn_frame, text="Doble clic en 'Solicitar' para cambiar cantidad", 
                font=("Segoe UI", 9), bg="white", fg="#64748b").pack(side="right", padx=10)
        
        form.center_on_parent = lambda: None
        x = (form.winfo_screenwidth() // 2) - (900 // 2)
        y = (form.winfo_screenheight() // 2) - (600 // 2)
        form.geometry(f"900x600+{x}+{y}")
        
        cargar_productos()
        
    def generar_planes(self):
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT i.id_inventario, p.id_producto, e.id_espacio, i.cant_actual, i.cant_maxima FROM inventarios i JOIN productos p ON i.id_producto = p.id_producto JOIN espacios_exhibicion e ON i.id_espacio = e.id_espacio WHERE i.cant_actual <= i.cant_minima")
        bajo_stock = cursor.fetchall()
        creados = 0
        for item in bajo_stock:
            cantidad = item["cant_maxima"] - item["cant_actual"]
            cursor.execute("SELECT COUNT(*) FROM planif_surtido WHERE id_producto=? AND id_espacio=? AND estado IN ('pendiente', 'aprobado')", (item["id_producto"], item["id_espacio"]))
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO planif_surtido (id_producto, id_espacio, fecha_planificada, cant_solicitada, estado) VALUES (?, ?, ?, ?, 'pendiente')",
                             (item["id_producto"], item["id_espacio"], datetime.now().strftime("%Y-%m-%d"), cantidad))
                creados += 1
        conn.commit()
        conn.close()
        self.cargar_surtido()
        messagebox.showinfo("Planes generados", f"Se generaron {creados} planes de surtido")
        
    def nuevo_cliente(self):
        form = tk.Toplevel(self.root)
        form.title("Nuevo Cliente")
        form.geometry("450x350")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["teal"])
        header.pack(fill="x")
        tk.Label(header, text="Nuevo Cliente", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["teal"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        campos = [("Nombre:", "nombre"), ("Apellido:", "apellido"), ("Documento:", "doc_identidad"), ("Telefono:", "telefono"), ("Email:", "email")]
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="w", pady=8)
            entry = ttk.Entry(card, width=35, font=("Segoe UI", 11))
            entry.grid(row=i, column=1, pady=8, padx=10)
            entries[key] = entry
            
        def guardar():
            if not entries["nombre"].get().strip():
                messagebox.showwarning("Error", "El nombre es obligatorio")
                return
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO clientes (nombre, apellido, doc_identidad, telefono, email) VALUES (?, ?, ?, ?, ?)",
                             (entries["nombre"].get(), entries["apellido"].get(), entries["doc_identidad"].get(), entries["telefono"].get(), entries["email"].get()))
                conn.commit()
                form.destroy()
                self.cargar_clientes()
                messagebox.showinfo("Exito", "Cliente guardado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=5, column=0, columnspan=2, pady=20)
        
    def editar_cliente(self):
        selection = self.cliente_tree.selection()
        if not selection:
            messagebox.showwarning("Seleccionar", "Seleccione un cliente")
            return
        item = self.cliente_tree.item(selection[0])
        values = item["values"]
        
        form = tk.Toplevel(self.root)
        form.title("Editar Cliente")
        form.geometry("450x350")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["teal"])
        header.pack(fill="x")
        tk.Label(header, text="Editar Cliente", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["teal"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM clientes WHERE id_cliente=?", (values[0],))
        datos = cursor.fetchone()
        conn.close()
        
        campos = [("Nombre:", "nombre"), ("Apellido:", "apellido"), ("Documento:", "doc_identidad"), ("Telefono:", "telefono"), ("Email:", "email")]
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white").grid(row=i, column=0, sticky="w", pady=8)
            entry = ttk.Entry(card, width=35, font=("Segoe UI", 11))
            entry.grid(row=i, column=1, pady=8, padx=10)
            entry.insert(0, datos[key] or "")
            entries[key] = entry
            
        def guardar():
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("UPDATE clientes SET nombre=?, apellido=?, doc_identidad=?, telefono=?, email=? WHERE id_cliente=?",
                             (entries["nombre"].get(), entries["apellido"].get(), entries["doc_identidad"].get(), entries["telefono"].get(), entries["email"].get(), values[0]))
                conn.commit()
                form.destroy()
                self.cargar_clientes()
                messagebox.showinfo("Exito", "Cliente actualizado")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Guardar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=5, column=0, columnspan=2, pady=20)
        
    def generar_liquidacion(self):
        form = tk.Toplevel(self.root)
        form.title("Generar Liquidacion")
        form.geometry("400x280")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        tk.Label(header, text="Generar Liquidacion", font=("Segoe UI", 14, "bold"), bg=self.S.COLORS["primary"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=25, pady=20)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT nombre FROM proveedores WHERE estado='activo'")
        proveedores = [p["nombre"] for p in cursor.fetchall()]
        conn.close()
        
        tk.Label(card, text="Proveedor:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=0, column=0, sticky="w", pady=10)
        prov_var = tk.StringVar()
        ttk.Combobox(card, textvariable=prov_var, values=proveedores, width=33, state="readonly").grid(row=0, column=1, pady=10, padx=10)
        
        tk.Label(card, text="Fecha Inicio:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=1, column=0, sticky="w", pady=10)
        entry_inicio = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        entry_inicio.grid(row=1, column=1, pady=10, padx=10)
        entry_inicio.insert(0, datetime.now().replace(day=1).strftime("%Y-%m-%d"))
        
        tk.Label(card, text="Fecha Fin:", font=("Segoe UI", 10, "bold"), bg="white").grid(row=2, column=0, sticky="w", pady=10)
        entry_fin = ttk.Entry(card, width=35, font=("Segoe UI", 11))
        entry_fin.grid(row=2, column=1, pady=10, padx=10)
        entry_fin.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        def guardar():
            if not prov_var.get():
                messagebox.showwarning("Error", "Seleccione un proveedor")
                return
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id_proveedor FROM proveedores WHERE nombre=?", (prov_var.get(),))
            id_prov = cursor.fetchone()["id_proveedor"]
            
            cursor.execute("SELECT id_contrato, porcentaje_comision FROM contratos WHERE id_proveedor=? AND estado='activo'", (id_prov,))
            contrato = cursor.fetchone()
            if not contrato:
                messagebox.showerror("Error", "No existe contrato activo")
                conn.close()
                return
                
            cursor.execute("""
                SELECT COALESCE(SUM(v.total), 0) as total FROM ventas v JOIN detalle_venta dv ON v.id_venta = dv.id_venta
                JOIN productos p ON dv.id_producto = p.id_proveedor WHERE p.id_proveedor = ? AND v.fecha_venta BETWEEN ? AND ? AND v.estado = 'completada'
            """, (id_prov, entry_inicio.get(), entry_fin.get() + " 23:59:59"))
            total_ventas = cursor.fetchone()["total"]
            
            if total_ventas == 0:
                messagebox.showinfo("Sin ventas", "No hay ventas en este periodo")
                conn.close()
                return
                
            monto_comision = total_ventas * (contrato["porcentaje_comision"] / 100)
            monto_pagar = total_ventas - monto_comision
            
            try:
                cursor.execute("INSERT INTO liquidaciones (id_proveedor, id_contrato, periodo_inicio, periodo_fin, total_ventas, monto_comision, monto_pagar, estado) VALUES (?, ?, ?, ?, ?, ?, ?, 'pendiente')",
                             (id_prov, contrato["id_contrato"], entry_inicio.get(), entry_fin.get(), total_ventas, monto_comision, monto_pagar))
                conn.commit()
                form.destroy()
                self.cargar_liquidaciones()
                messagebox.showinfo("Exito", "Liquidacion generada")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                conn.close()
                
        tk.Button(card, text="Generar", font=("Segoe UI", 11, "bold"), bg=self.S.COLORS["success"], fg="white", relief="flat", cursor="hand2", padx=30, pady=10, command=guardar).grid(row=3, column=0, columnspan=2, pady=25)
        
    # ==================== METODOS POS ====================
    
    def abrir_modal_productos(self):
        modal = tk.Toplevel(self.root)
        modal.title("Catalogo de Productos")
        modal.geometry("700x550")
        modal.configure(bg="white")
        modal.transient(self.root)
        modal.grab_set()
        
        header = tk.Frame(modal, bg="#107c10")
        header.pack(fill="x")
        
        tk.Label(header, text="   CATALOGO DE PRODUCTOS", font=("Segoe UI", 14, "bold"),
                bg="#107c10", fg="white").pack(pady=12, anchor="w", padx=15)
        
        search_frame = tk.Frame(modal, bg="white")
        search_frame.pack(fill="x", padx=20, pady=15)
        
        tk.Label(search_frame, text="Buscar:", font=("Segoe UI", 11), bg="white").pack(side="left", padx=(0, 10))
        
        modal_entry = ttk.Entry(search_frame, width=50, font=("Segoe UI", 12))
        modal_entry.pack(side="left", ipady=6, fill="x", expand=True)
        
        table_frame = tk.Frame(modal, bg="white")
        table_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        columns = ("Codigo", "Producto", "Precio", "Imp", "Stock")
        modal_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        modal_tree.heading("Codigo", text="CODIGO")
        modal_tree.heading("Producto", text="PRODUCTO")
        modal_tree.heading("Precio", text="PRECIO")
        modal_tree.heading("Imp", text="IMP")
        modal_tree.heading("Stock", text="STOCK")
        
        modal_tree.column("Codigo", width=100, anchor="center")
        modal_tree.column("Producto", width=280)
        modal_tree.column("Precio", width=100, anchor="e")
        modal_tree.column("Imp", width=50, anchor="center")
        modal_tree.column("Stock", width=70, anchor="center")
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=modal_tree.yview)
        modal_tree.configure(yscrollcommand=scrollbar.set)
        modal_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        modal_productos = []
        curr = self.get_currency_info()
        
        def cargar_productos_modal():
            for item in modal_tree.get_children():
                modal_tree.delete(item)
            
            conn = get_connection()
            cursor = conn.cursor()
            texto = modal_entry.get().strip()
            
            if texto:
                cursor.execute("""
                    SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, SUM(i.cant_actual) as cant_actual, i.id_inventario, i.id_producto, i.id_espacio
                    FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
                    WHERE p.estado = 'activo' AND i.cant_actual > 0 AND (p.nombre LIKE ? OR p.codigo_barras LIKE ?)
                    GROUP BY p.id_producto ORDER BY p.nombre
                """, (f"%{texto}%", f"%{texto}%"))
            else:
                cursor.execute("""
                    SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, SUM(i.cant_actual) as cant_actual, i.id_inventario, i.id_producto, i.id_espacio
                    FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
                    WHERE p.estado = 'activo' AND i.cant_actual > 0
                    GROUP BY p.id_producto ORDER BY p.nombre
                """)
            
            modal_productos.clear()
            for row in cursor.fetchall():
                row_dict = dict(row)
                precio = f"{curr['simbolo_local']} {row_dict['precio_venta']:,.2f}"
                tipo_imp = row_dict.get("tipo_impuesto", "gravado")
                tipo_label = {"gravado": "G", "exento": "E", "zero": "Z"}.get(tipo_imp, "G")
                modal_tree.insert("", "end", values=(row_dict["codigo_barras"], row_dict["nombre"], precio, tipo_label, row_dict["cant_actual"]))
                modal_productos.append(row_dict)
            conn.close()
        
        def agregar_desde_modal():
            selection = modal_tree.selection()
            if not selection:
                messagebox.showwarning("Seleccionar", "Seleccione un producto")
                return
            
            item = modal_tree.item(selection[0])
            values = item["values"]
            nombre = values[1]
            
            for prod in modal_productos:
                if prod["nombre"] == nombre:
                    for cart_item in self.carrito:
                        if cart_item["nombre"] == nombre:
                            if cart_item["stock"] > cart_item["cantidad"]:
                                cart_item["cantidad"] += 1
                                self.actualizar_carrito()
                            modal.destroy()
                            return
                    
                    tipo_imp = prod.get("tipo_impuesto", "gravado")
                    self.carrito.append({
                        "nombre": nombre, "precio": prod["precio_venta"], "stock": prod["cant_actual"],
                        "cantidad": 1, "id_inventario": prod["id_inventario"], "id_producto": prod["id_producto"], 
                        "id_espacio": prod["id_espacio"], "tipo_impuesto": tipo_imp
                    })
                    self.actualizar_carrito()
                    modal.destroy()
                    return
        
        modal_entry.bind("<KeyRelease>", lambda e: cargar_productos_modal())
        modal_tree.bind("<Double-1>", lambda e: agregar_desde_modal())
        modal_tree.bind("<Return>", lambda e: agregar_desde_modal())
        
        btn_frame = tk.Frame(modal, bg="white")
        btn_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        tk.Button(btn_frame, text="Agregar Seleccionado", font=("Segoe UI", 12, "bold"),
                 bg="#107c10", fg="white", relief="flat", cursor="hand2",
                 command=agregar_desde_modal, padx=20, pady=10).pack(side="left")
        
        tk.Button(btn_frame, text="Cerrar", font=("Segoe UI", 11),
                 bg="#605e5c", fg="white", relief="flat", cursor="hand2",
                 command=modal.destroy, padx=20, pady=10).pack(side="right")
        
        cargar_productos_modal()
        
    def agregar_producto_rapido(self):
        texto = self.pos_entry.get().strip()
        if not texto or len(texto) < 2:
            return
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.codigo_barras, p.nombre, p.precio_venta, p.tipo_impuesto, SUM(i.cant_actual) as cant_actual, i.id_inventario, i.id_producto, i.id_espacio
            FROM productos p JOIN inventarios i ON p.id_producto = i.id_producto
            WHERE p.estado = 'activo' AND i.cant_actual > 0 AND (p.nombre LIKE ? OR p.codigo_barras LIKE ?)
            GROUP BY p.id_producto ORDER BY p.nombre LIMIT 10
        """, (f"%{texto}%", f"%{texto}%"))
        productos = cursor.fetchall()
        conn.close()
        
        if not productos:
            messagebox.showinfo("No encontrado", "Producto no encontrado o sin stock")
            return
        
        if len(productos) == 1:
            prod = dict(productos[0])
            self._agregar_al_carrito(prod)
            self.pos_entry.delete(0, tk.END)
            return
        
        form_seleccion = tk.Toplevel(self.root)
        form_seleccion.title("Seleccionar Producto")
        form_seleccion.geometry("450x350")
        form_seleccion.configure(bg="white")
        form_seleccion.transient(self.root)
        form_seleccion.grab_set()
        
        tk.Label(form_seleccion, text="PRODUCTOS ENCONTRADOS", font=("Segoe UI", 14, "bold"),
                bg="white", fg="#323130").pack(pady=15)
        
        frame_lista = tk.Frame(form_seleccion, bg="white")
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)
        
        lista_productos = tk.Listbox(frame_lista, font=("Segoe UI", 11), height=10)
        lista_productos.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(frame_lista, orient="vertical")
        scrollbar.config(command=lista_productos.yview)
        scrollbar.pack(side="right", fill="y")
        lista_productos.config(yscrollcommand=scrollbar.set)
        
        productos_dict = {}
        for i, prod in enumerate(productos):
            prod_dict = dict(prod)
            precio = f"{curr['simbolo_local']} {prod_dict['precio_venta']:,.2f}"
            texto_lista = f"{prod_dict['nombre'][:35]} | {precio} | Stock: {prod_dict['cant_actual']}"
            lista_productos.insert(tk.END, texto_lista)
            productos_dict[i] = prod_dict
        
        def seleccionar_producto(event=None):
            seleccion = lista_productos.curselection()
            if not seleccion:
                return
            idx = int(seleccion[0])
            prod = productos_dict[idx]
            self._agregar_al_carrito(prod)
            form_seleccion.destroy()
            self.pos_entry.delete(0, tk.END)
        
        lista_productos.bind("<Double-Button-1>", seleccionar_producto)
        lista_productos.bind("<Return>", seleccionar_producto)
        
        frame_botones = tk.Frame(form_seleccion, bg="white")
        frame_botones.pack(pady=10)
        
        tk.Button(frame_botones, text="Agregar Seleccionado", font=("Segoe UI", 11, "bold"),
                 bg="#107c10", fg="white", relief="flat", command=seleccionar_producto,
                 padx=20, pady=8).pack(side="left", padx=5)
        
        tk.Button(frame_botones, text="Cancelar", font=("Segoe UI", 10),
                 bg="#605e5c", fg="white", relief="flat", command=form_seleccion.destroy,
                 padx=20, pady=8).pack(side="left", padx=5)
        
        lista_productos.focus()
        lista_productos.select_set(0)
    
    def _agregar_al_carrito(self, prod):
        for cart_item in self.carrito:
            if cart_item["nombre"] == prod["nombre"]:
                if cart_item["stock"] > cart_item["cantidad"]:
                    cart_item["cantidad"] += 1
                    self.actualizar_carrito()
                return
        
        tipo_impuesto = prod.get("tipo_impuesto", "gravado")
        self.carrito.append({
            "nombre": prod["nombre"], "precio": prod["precio_venta"], "stock": prod["cant_actual"],
            "cantidad": 1, "id_inventario": prod["id_inventario"], "id_producto": prod["id_producto"], 
            "id_espacio": prod["id_espacio"], "tipo_impuesto": tipo_impuesto
        })
        self.actualizar_carrito()
            
    def agregar_al_carrito(self):
        if self.pos_tree is None:
            return
        selection = self.pos_tree.selection()
        if not selection:
            return
        item = self.pos_tree.item(selection[0])
        values = item["values"]
        
        nombre = values[1]
        
        for prod in self.pos_productos:
            if prod["nombre"] == nombre:
                for cart_item in self.carrito:
                    if cart_item["nombre"] == nombre:
                        if cart_item["stock"] > cart_item["cantidad"]:
                            cart_item["cantidad"] += 1
                            self.actualizar_carrito()
                        return
                
                tipo_impuesto = prod.get("tipo_impuesto", "gravado")
                self.carrito.append({
                    "nombre": nombre, "precio": prod["precio_venta"], "stock": prod["cant_actual"],
                    "cantidad": 1, "id_inventario": prod["id_inventario"], "id_producto": prod["id_producto"], 
                    "id_espacio": prod["id_espacio"], "tipo_impuesto": tipo_impuesto
                })
                self.actualizar_carrito()
                return
                
    def actualizar_carrito(self):
        curr = self.get_currency_info()
        
        if self.cart_tree is None:
            return
            
        for item in self.cart_tree.get_children():
            self.cart_tree.delete(item)
        
        subtotal = 0
        total_items = 0
        total_gravado = 0
        total_exento = 0
        total_zero = 0
        
        for item in self.carrito:
            item_subtotal = item["precio"] * item["cantidad"]
            tipo_imp = item.get("tipo_impuesto", "gravado")
            
            tipo_label = {"gravado": "G", "exento": "E", "zero": "Z"}.get(tipo_imp, "G")
            
            self.cart_tree.insert("", "end", values=(
                item["nombre"][:35],
                f"{curr['simbolo_local']} {item['precio']:,.2f}",
                item["cantidad"],
                tipo_label,
                f"{curr['simbolo_local']} {item_subtotal:,.2f}"
            ))
            
            subtotal += item_subtotal
            total_items += item["cantidad"]
            
            if tipo_imp == "gravado":
                total_gravado += item_subtotal
            elif tipo_imp == "exento":
                total_exento += item_subtotal
            elif tipo_imp == "zero":
                total_zero += item_subtotal
        
        impuesto = total_gravado * 0.15
        total = subtotal + impuesto
        
        if hasattr(self, 'cart_count_label') and self.cart_count_label:
            self.cart_count_label.config(text=f"{total_items} items")
        
        if self.subtotal_label:
            self.subtotal_label.config(text=f"{curr['simbolo_local']} {subtotal:,.2f}")
        if self.impuesto_label:
            self.impuesto_label.config(text=f"{curr['simbolo_local']} {impuesto:,.2f} (G:{curr['simbolo_local']} {total_gravado:,.2f})")
        if self.total_label:
            self.total_label.config(text=f"{curr['simbolo_local']} {total:,.2f}")
        if self.total_usd_label:
            self.total_usd_label.config(text=f"{curr['simbolo_extranjera']} {total / curr['tasa']:,.2f}")
        
    def aumentar_cantidad(self):
        if self.cart_tree is None:
            return
        selection = self.cart_tree.selection()
        if not selection:
            return
        item = self.cart_tree.item(selection[0])
        values = item["values"]
        for cart_item in self.carrito:
            if cart_item["nombre"] == values[0]:
                if cart_item["stock"] > cart_item["cantidad"]:
                    cart_item["cantidad"] += 1
                    self.actualizar_carrito()
                return
                
    def disminuir_cantidad(self):
        if self.cart_tree is None:
            return
        selection = self.cart_tree.selection()
        if not selection:
            return
        item = self.cart_tree.item(selection[0])
        values = item["values"]
        for i, cart_item in enumerate(self.carrito):
            if cart_item["nombre"] == values[0]:
                if cart_item["cantidad"] > 1:
                    cart_item["cantidad"] -= 1
                else:
                    self.carrito.pop(i)
                self.actualizar_carrito()
                return
                
    def eliminar_carrito(self):
        if self.cart_tree is None:
            return
        selection = self.cart_tree.selection()
        if not selection:
            return
        item = self.cart_tree.item(selection[0])
        values = item["values"]
        for i, cart_item in enumerate(self.carrito):
            if cart_item["nombre"] == values[0]:
                self.carrito.pop(i)
                self.actualizar_carrito()
                return
                
    def vaciar_carrito(self):
        if not self.carrito:
            return
        if messagebox.askyesno("Vaciar Carrito", "Esta seguro de vaciar el carrito?"):
            self.carrito = []
            self.actualizar_carrito()
                
    def editar_cantidad_carrito(self):
        selection = self.cart_tree.selection()
        if not selection:
            return
        item = self.cart_tree.item(selection[0])
        values = item["values"]
        
        for cart_item in self.carrito:
            if cart_item["nombre"] == values[0]:
                dialog = tk.Toplevel(self.root)
                dialog.title("Editar Cantidad")
                dialog.geometry("450x350")
                dialog.configure(bg="#f8fafc")
                dialog.transient(self.root)
                dialog.grab_set()
                
                screen_width = dialog.winfo_screenwidth()
                screen_height = dialog.winfo_screenheight()
                x = (screen_width // 2) - (450 // 2)
                y = (screen_height // 2) - (350 // 2)
                dialog.geometry(f"450x350+{x}+{y}")
                
                header = tk.Frame(dialog, bg="#1e3a5f", height=50)
                header.pack(fill="x")
                header.pack_propagate(False)
                tk.Label(header, text="EDITAR CANTIDAD", font=("Segoe UI", 14, "bold"),
                        bg="#1e3a5f", fg="white").pack(pady=12)
                
                card = tk.Frame(dialog, bg="white")
                card.pack(fill="both", expand=True, padx=30, pady=20)
                
                nombre_producto = cart_item['nombre'][:40] + "..." if len(cart_item['nombre']) > 40 else cart_item['nombre']
                tk.Label(card, text=f"Producto:", font=("Segoe UI", 11), bg="white", fg="#64748b").pack(anchor="w")
                tk.Label(card, text=nombre_producto, font=("Segoe UI", 13, "bold"), bg="white", fg="#1e293b").pack(anchor="w", pady=(0, 10))
                
                tk.Label(card, text=f"Stock disponible: {cart_item['stock']} unidades", font=("Segoe UI", 10), bg="white", fg="#64748b").pack(anchor="w", pady=(0, 15))
                
                tk.Label(card, text="Nueva cantidad:", font=("Segoe UI", 11), bg="white", fg="#475569").pack(anchor="w", pady=(5, 5))
                
                cantidad_var = tk.IntVar(value=cart_item["cantidad"])
                spinbox = ttk.Spinbox(card, from_=1, to=cart_item["stock"], textvariable=cantidad_var, width=25, font=("Segoe UI", 18))
                spinbox.pack(pady=10, ipady=8)
                spinbox.focus()
                
                btn_frame = tk.Frame(card, bg="white")
                btn_frame.pack(fill="x", pady=(25, 0))
                
                def guardar():
                    cart_item["cantidad"] = cantidad_var.get()
                    self.actualizar_carrito()
                    dialog.destroy()
                
                tk.Button(btn_frame, text="Cancelar", font=("Segoe UI", 12), bg="#e2e8f0", fg="#475569",
                         relief="flat", command=dialog.destroy, padx=20, pady=10).pack(side="left", fill="x", expand=True, padx=(0, 10))
                tk.Button(btn_frame, text="Guardar", font=("Segoe UI", 13, "bold"), bg=self.S.COLORS["success"], fg="white",
                         relief="flat", command=guardar, padx=20, pady=10).pack(side="right", fill="x", expand=True)
                
                spinbox.bind("<Return>", lambda e: guardar())
                return
                
    def nueva_venta_pos(self):
        if self.carrito:
            if messagebox.askyesno("Nueva Venta", "Hay productos en el carrito. Desea descartarlos?"):
                self.carrito = []
                self.actualizar_carrito()
            else:
                return
        self.pos_entry.focus()
                
    def procesar_venta(self):
        if not self.carrito:
            messagebox.showwarning("Carrito vacio", "Agregue productos al carrito")
            return
            
        if not hasattr(self, 'pos_pago'):
            messagebox.showwarning("Error", "El sistema no esta listo. Recargue la pantalla.")
            return
            
        metodo_pago = self.pos_pago.get()
        if not metodo_pago:
            self.pos_pago.current(0)
            metodo_pago = self.pos_pago.get()
            
        referencia = self.pos_referencia.get().strip()
        cliente = self.pos_cliente.get()
        
        curr = self.get_currency_info()
        subtotal = sum(item["precio"] * item["cantidad"] for item in self.carrito)
        impuesto = subtotal * 0.15
        total = subtotal + impuesto
        
        efectivo_recibido = 0
        cambio = 0
        
        if metodo_pago == "Efectivo":
            form_pago = tk.Toplevel(self.root)
            form_pago.title("Pago en Efectivo")
            form_pago.geometry("450x320")
            form_pago.configure(bg="#f8fafc")
            form_pago.transient(self.root)
            form_pago.grab_set()
            
            screen_width = form_pago.winfo_screenwidth()
            screen_height = form_pago.winfo_screenheight()
            x = (screen_width // 2) - (450 // 2)
            y = (screen_height // 2) - (320 // 2)
            form_pago.geometry(f"450x320+{x}+{y}")
            
            header = tk.Frame(form_pago, bg="#1e3a5f", height=60)
            header.pack(fill="x")
            header.pack_propagate(False)
            tk.Label(header, text="PAGO EN EFECTIVO", font=("Segoe UI", 18, "bold"),
                    bg="#1e3a5f", fg="white").pack(pady=15)
            
            card = tk.Frame(form_pago, bg="white")
            card.pack(fill="both", expand=True, padx=30, pady=20)
            
            total_frame = tk.Frame(card, bg="#ecfdf5", padx=20, pady=15)
            total_frame.pack(fill="x", pady=(0, 20))
            tk.Label(total_frame, text="TOTAL A PAGAR", font=("Segoe UI", 12, "bold"),
                    bg="#ecfdf5", fg="#059669").pack()
            tk.Label(total_frame, text=f"{curr['simbolo_local']} {total:,.2f}",
                    font=("Segoe UI", 32, "bold"), bg="#ecfdf5", fg="#059669").pack()
            
            tk.Label(card, text="Dinero recibido del cliente:", font=("Segoe UI", 12),
                    bg="white", fg="#475569").pack(anchor="w", pady=(0, 5))
            
            entry_efectivo = ttk.Entry(card, width=30, font=("Segoe UI", 18), justify="center")
            entry_efectivo.pack(pady=5, ipady=10)
            entry_efectivo.focus()
            
            label_cambio = tk.Label(card, text="Cambio: L 0.00",
                    font=("Segoe UI", 20, "bold"), bg="white", fg="#64748b")
            label_cambio.pack(pady=15)
            
            def calcular_cambio(*args):
                try:
                    efectivo = float(entry_efectivo.get() or 0)
                    if efectivo >= total:
                        cambio_val = efectivo - total
                        label_cambio.config(text=f"Cambio: {curr['simbolo_local']} {cambio_val:,.2f}", fg="#059669")
                        btn_confirmar.config(state="normal")
                    else:
                        label_cambio.config(text=f"Falta: {curr['simbolo_local']} {total - efectivo:,.2f}", fg="#dc2626")
                        btn_confirmar.config(state="disabled")
                except:
                    label_cambio.config(text="Cambio: L 0.00", fg="#64748b")
                    btn_confirmar.config(state="disabled")
            
            entry_efectivo.bind("<KeyRelease>", calcular_cambio)
            
            btn_frame = tk.Frame(card, bg="white")
            btn_frame.pack(fill="x", pady=(10, 0))
            
            btn_cancelar = tk.Button(btn_frame, text="Cancelar", font=("Segoe UI", 12),
                                    bg="#e2e8f0", fg="#475569", relief="flat",
                                    command=form_pago.destroy, padx=20, pady=12)
            btn_cancelar.pack(side="left", padx=(0, 10), fill="x", expand=True)
            
            btn_confirmar = tk.Button(btn_frame, text="Confirmar Pago", font=("Segoe UI", 14, "bold"),
                                     bg="#059669", fg="white", relief="flat", state="disabled",
                                     padx=20, pady=12)
            btn_confirmar.pack(side="right", fill="x", expand=True)
            
            def confirmar_desde_form():
                try:
                    nonlocal efectivo_recibido, cambio
                    efectivo_recibido = float(entry_efectivo.get())
                    cambio = efectivo_recibido - total
                    form_pago.destroy()
                    self.completar_venta(total, subtotal, impuesto, metodo_pago, referencia, cliente, efectivo_recibido, cambio)
                except:
                    pass
            
            btn_confirmar.config(command=confirmar_desde_form)
            entry_efectivo.bind("<Return>", lambda e: confirmar_desde_form())
            return
        
        self.completar_venta(total, subtotal, impuesto, metodo_pago, referencia, cliente, efectivo_recibido, cambio)
    
    def completar_venta(self, total, subtotal, impuesto, metodo_pago, referencia, cliente_nombre, efectivo_recibido=0, cambio=0):
        curr = self.get_currency_info()
        
        conn = get_connection()
        cursor = conn.cursor()
        try:
            id_cliente = None
            if cliente_nombre and cliente_nombre != "Consumidor Final":
                cursor.execute("SELECT id_cliente FROM clientes WHERE nombre || ' ' || COALESCE(apellido, '') = ?", (cliente_nombre,))
                result = cursor.fetchone()
                if result:
                    id_cliente = result["id_cliente"]
            
            referencia_completa = referencia
            if efectivo_recibido > 0:
                referencia_completa = f"{referencia} | Efectivo: {curr['simbolo_local']} {efectivo_recibido:,.2f} | Cambio: {curr['simbolo_local']} {cambio:,.2f}" if referencia else f"Efectivo: {curr['simbolo_local']} {efectivo_recibido:,.2f} | Cambio: {curr['simbolo_local']} {cambio:,.2f}"
            
            cursor.execute("""
                INSERT INTO ventas (id_cliente, id_usuario, total, subtotal, impuesto, metodo_pago, estado, referencia)
                VALUES (?, ?, ?, ?, ?, ?, 'completada', ?)
            """, (id_cliente, self.usuario_actual["id_usuario"], total, subtotal, impuesto, metodo_pago, referencia_completa))
            id_venta = cursor.lastrowid
            
            productos_vendidos = []
            for item in self.carrito:
                tipo_imp = item.get("tipo_impuesto", "gravado")
                cursor.execute("""
                    INSERT INTO detalle_venta (id_venta, id_producto, id_espacio, cantidad, precio_unit, subtotal, tipo_impuesto)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (id_venta, item["id_producto"], item["id_espacio"], item["cantidad"], item["precio"], item["precio"] * item["cantidad"], tipo_imp))
                
                cursor.execute("""
                    UPDATE inventarios SET cant_actual = cant_actual - ?
                    WHERE id_inventario = ?
                """, (item["cantidad"], item["id_inventario"]))
                
                productos_vendidos.append({
                    "nombre": item["nombre"],
                    "cantidad": item["cantidad"],
                    "precio": item["precio"],
                    "subtotal": item["precio"] * item["cantidad"],
                    "tipo_impuesto": tipo_imp
                })
            
            conn.commit()
            
            self.generar_ticket_termico(id_venta, productos_vendidos, subtotal, impuesto, total, metodo_pago, cliente_nombre, efectivo_recibido, cambio)
            
            self.carrito = []
            self.actualizar_carrito()
            self.pos_entry.delete(0, tk.END)
            
            if cambio > 0:
                messagebox.showinfo("Venta Exitosa", f"Venta #{id_venta} procesada!\n\nTotal: {curr['simbolo_local']} {total:,.2f}\nEfectivo: {curr['simbolo_local']} {efectivo_recibido:,.2f}\nCambio: {curr['simbolo_local']} {cambio:,.2f}")
            else:
                messagebox.showinfo("Venta Exitosa", f"Venta #{id_venta} procesada correctamente!\nTotal: {curr['simbolo_local']} {total:,.2f}")
            
        except Exception as e:
            conn.rollback()
            messagebox.showerror("Error", str(e))
        finally:
            conn.close()
            
    def generar_ticket_termico(self, id_venta, productos, subtotal, impuesto, total, metodo_pago, cliente, efectivo_recibido=0, cambio=0):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            import datetime
            
            ancho_pagina = 80 * mm
            alto_pagina = 200 * mm + (len(productos) * 8 * mm)
            
            c = canvas.Canvas(f"ticket_{id_venta}.pdf", pagesize=(ancho_pagina, alto_pagina))
            
            config = get_config_sistema()
            curr = self.get_currency_info()
            
            y = alto_pagina - 15 * mm
            c.setFont("Helvetica-Bold", 12)
            c.drawCentredString(ancho_pagina / 2, y, config.get("nombre_tienda", "TIENDA CONCEPTO"))
            
            y -= 6 * mm
            c.setFont("Helvetica", 8)
            c.drawCentredString(ancho_pagina / 2, y, "RNC/NIT: 000-00000-0")
            c.drawCentredString(ancho_pagina / 2, y - 4 * mm, "Tel: 0000-0000")
            
            y -= 12 * mm
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(ancho_pagina / 2, y, f"FACTURA #{id_venta}")
            
            y -= 6 * mm
            c.setFont("Helvetica", 8)
            fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
            c.drawCentredString(ancho_pagina / 2, y, fecha)
            c.drawCentredString(ancho_pagina / 2, y - 4 * mm, f"Atendido: {self.usuario_actual['nombre']}")
            
            y -= 10 * mm
            c.line(10 * mm, y, ancho_pagina - 10 * mm, y)
            
            y -= 5 * mm
            c.setFont("Helvetica", 7)
            c.drawString(10 * mm, y, "CLIENTE:")
            c.setFont("Helvetica-Bold", 8)
            c.drawString(30 * mm, y, cliente)
            
            y -= 5 * mm
            c.setFont("Helvetica", 7)
            c.drawString(10 * mm, y, "PAGO:")
            c.setFont("Helvetica-Bold", 8)
            c.drawString(30 * mm, y, metodo_pago)
            
            y -= 8 * mm
            c.line(10 * mm, y, ancho_pagina - 10 * mm, y)
            
            c.setFont("Helvetica-Bold", 8)
            c.drawString(10 * mm, y - 5 * mm, "PRODUCTO")
            c.drawRightString(ancho_pagina - 10 * mm, y - 5 * mm, "SUBTOTAL")
            
            y -= 7 * mm
            c.setFont("Helvetica", 7)
            
            for prod in productos:
                if y < 30 * mm:
                    c.showPage()
                    y = alto_pagina - 15 * mm
                
                nombre = prod["nombre"][:20] if len(prod["nombre"]) > 20 else prod["nombre"]
                precio_str = f"{curr['simbolo_local']} {prod['subtotal']:,.2f}"
                linea = f"{prod['cantidad']} x {nombre}"
                c.drawString(10 * mm, y, linea)
                c.drawRightString(ancho_pagina - 10 * mm, y, precio_str)
                y -= 5 * mm
            
            y -= 3 * mm
            c.line(10 * mm, y, ancho_pagina - 10 * mm, y)
            
            c.setFont("Helvetica", 8)
            c.drawString(10 * mm, y - 5 * mm, "Subtotal:")
            c.drawRightString(ancho_pagina - 10 * mm, y - 5 * mm, f"{curr['simbolo_local']} {subtotal:,.2f}")
            
            c.drawString(10 * mm, y - 10 * mm, "Impuesto (15%):")
            c.drawRightString(ancho_pagina - 10 * mm, y - 10 * mm, f"{curr['simbolo_local']} {impuesto:,.2f}")
            
            y -= 17 * mm
            c.setFont("Helvetica-Bold", 10)
            c.drawString(10 * mm, y, "TOTAL:")
            c.drawRightString(ancho_pagina - 10 * mm, y, f"{curr['simbolo_local']} {total:,.2f}")
            
            y -= 6 * mm
            c.setFont("Helvetica", 8)
            c.drawRightString(ancho_pagina - 10 * mm, y, f"(={curr['simbolo_extranjera']} {total / curr['tasa']:,.2f})")
            
            if efectivo_recibido > 0:
                y -= 7 * mm
                c.setFont("Helvetica", 7)
                c.drawString(10 * mm, y, "Efectivo:")
                c.drawRightString(ancho_pagina - 10 * mm, y, f"{curr['simbolo_local']} {efectivo_recibido:,.2f}")
                y -= 5 * mm
                c.setFont("Helvetica-Bold", 8)
                c.drawString(10 * mm, y, "Cambio:")
                c.drawRightString(ancho_pagina - 10 * mm, y, f"{curr['simbolo_local']} {cambio:,.2f}")
                y -= 8 * mm
            else:
                y -= 8 * mm
            
            c.setFont("Helvetica", 7)
            c.drawCentredString(ancho_pagina / 2, y, "GRACIAS POR SU COMPRA")
            c.drawCentredString(ancho_pagina / 2, y - 4 * mm, "Este documento es valido como factura")
            
            c.save()
            
            import os
            ticket_path = os.path.abspath(f"ticket_{id_venta}.pdf")
            messagebox.showinfo("Ticket Generado", f"Venta #{id_venta} procesada!\n\nTicket guardado en:\n{ticket_path}")
            
        except ImportError:
            messagebox.showinfo("Ticket", f"Venta #{id_venta} procesada!\nNo se pudo generar PDF (reportlab no instalado)")
        except Exception as e:
            messagebox.showinfo("Ticket", f"Venta #{id_venta} procesada!\nError generando PDF: {str(e)}")
            
    def gestionar_devolucion(self):
        form = tk.Toplevel(self.root)
        form.title("Gestionar Devolucion")
        form.geometry("800x600")
        form.configure(bg="#f1f5f9")
        form.transient(self.root)
        form.grab_set()
        
        header = tk.Frame(form, bg="#dc2626")
        header.pack(fill="x")
        tk.Label(header, text="DEVOLUCIONES", font=("Segoe UI", 16, "bold"),
                bg="#dc2626", fg="white").pack(pady=12, padx=20)
        
        search_frame = tk.Frame(form, bg="white")
        search_frame.pack(fill="x", padx=20, pady=15)
        
        tk.Label(search_frame, text="Buscar Venta (#):", font=("Segoe UI", 11), bg="white").pack(side="left", padx=5)
        entry_busqueda = ttk.Entry(search_frame, width=15, font=("Segoe UI", 12))
        entry_busqueda.pack(side="left", padx=5)
        
        tk.Button(search_frame, text="Buscar", font=("Segoe UI", 11, "bold"), bg="#3b82f6", fg="white",
                 relief="flat", command=lambda: buscar_venta(entry_busqueda.get().strip())).pack(side="left", padx=10)
        
        resultado_frame = tk.Frame(form, bg="white")
        resultado_frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        tk.Label(resultado_frame, text="Detalle de la Venta:", font=("Segoe UI", 11, "bold"), bg="white").pack(anchor="w")
        
        columns = ("Producto", "Cantidad", "Precio", "Subtotal")
        tree_dev = ttk.Treeview(resultado_frame, columns=columns, show="headings", height=10)
        for col in columns:
            tree_dev.heading(col, text=col)
            tree_dev.column(col, width=150 if col != "Producto" else 350)
        tree_dev.pack(fill="both", expand=True, pady=10)
        
        info_label = tk.Label(resultado_frame, text="Ingrese el numero de venta y haga clic en Buscar", 
                            font=("Segoe UI", 11), bg="white", fg="#64748b")
        info_label.pack(anchor="w")
        
        venta_actual = {"id": None}
        
        def buscar_venta(texto):
            for item in tree_dev.get_children():
                tree_dev.delete(item)
            
            if not texto:
                messagebox.showwarning("Buscar", "Ingrese numero de venta")
                return
            
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT v.*, u.nombre as vendedor FROM ventas v
                JOIN usuarios u ON v.id_usuario = u.id_usuario
                WHERE v.id_venta = ? AND v.estado = 'completada'
            """, (texto,))
            venta = cursor.fetchone()
            
            if not venta:
                info_label.config(text="Venta no encontrada o ya fue devuelta/anulada")
                info_label.config(fg="#dc2626")
                conn.close()
                return
            
            cursor.execute("""
                SELECT dv.*, p.nombre as producto FROM detalle_venta dv
                JOIN productos p ON dv.id_producto = p.id_producto
                WHERE dv.id_venta = ?
            """, (texto,))
            detalles = cursor.fetchall()
            conn.close()
            
            venta_actual["id"] = venta["id_venta"]
            venta_actual["fecha"] = venta["fecha_venta"]
            venta_actual["metodo"] = venta["metodo_pago"]
            venta_actual["detalles"] = detalles
            
            total_dev = 0
            for det in detalles:
                tree_dev.insert("", "end", values=(det["producto"], det["cantidad"], f"L {det['precio_unit']:.2f}", f"L {det['subtotal']:.2f}"))
                total_dev += det["subtotal"]
            
            info_label.config(text=f"Venta #{venta['id_venta']} | Fecha: {venta['fecha_venta'][:10]} | Total: L {total_dev:,.2f} | Metodo: {venta['metodo_pago']}")
            info_label.config(fg="#1e293b")
        
        def procesar_devolucion():
            selection = tree_dev.selection()
            if not selection:
                messagebox.showwarning("Seleccionar", "Seleccione un producto para devolver")
                return
            
            if not venta_actual["id"]:
                messagebox.showwarning("Buscar", "Primero busque una venta")
                return
            
            item = tree_dev.item(selection[0])
            producto_nombre = item["values"][0]
            cantidad_original = item["values"][1]
            
            dialog = tk.Toplevel(form)
            dialog.title("Realizar Devolucion")
            dialog.geometry("400x280")
            dialog.configure(bg="white")
            dialog.transient(form)
            dialog.grab_set()
            
            tk.Label(dialog, text="REALIZAR DEVOLUCION", font=("Segoe UI", 14, "bold"),
                    bg="#dc2626", fg="white").pack(fill="x", pady=0)
            
            inner = tk.Frame(dialog, bg="white")
            inner.pack(fill="both", expand=True, padx=20, pady=20)
            
            tk.Label(inner, text=f"Producto: {producto_nombre}", font=("Segoe UI", 11, "bold"), bg="white").pack(anchor="w")
            tk.Label(inner, text=f"Cantidad original: {cantidad_original}", font=("Segoe UI", 10), bg="white").pack(anchor="w", pady=(5, 15))
            
            tk.Label(inner, text="Cantidad a devolver:", font=("Segoe UI", 11), bg="white").pack(anchor="w")
            spin_cant = ttk.Spinbox(inner, from_=1, to=cantidad_original, width=25, font=("Segoe UI", 14))
            spin_cant.pack(pady=(0, 15))
            spin_cant.delete(0, tk.END)
            spin_cant.insert(0, "1")
            
            tk.Label(inner, text="Motivo:", font=("Segoe UI", 11), bg="white").pack(anchor="w")
            combo_motivo = ttk.Combobox(inner, values=["Defectuoso", "No era lo esperado", "Cambio de opinion", "Error en pedido", "Otro"], width=30, font=("Segoe UI", 11))
            combo_motivo.current(0)
            combo_motivo.pack(pady=(0, 15))
            
            def confirmar_devolucion():
                cantidad = int(spin_cant.get())
                motivo = combo_motivo.get()
                id_venta = venta_actual["id"]
                
                conn = get_connection()
                cursor = conn.cursor()
                try:
                    cursor.execute("""
                        INSERT INTO devoluciones (id_venta, id_usuario, cantidad, motivo, estado, fecha_devolucion)
                        VALUES (?, ?, ?, ?, 'completada', datetime('now'))
                    """, (id_venta, self.usuario_actual["id_usuario"], cantidad, motivo))
                    
                    cursor.execute("""
                        UPDATE ventas SET estado = 'devuelta' WHERE id_venta = ?
                    """, (id_venta,))
                    
                    cursor.execute("""
                        SELECT id_producto, id_espacio FROM detalle_venta 
                        WHERE id_venta = ? AND id_producto = (
                            SELECT id_producto FROM productos WHERE nombre = ?
                        )
                    """, (id_venta, producto_nombre))
                    detalle = cursor.fetchone()
                    
                    if detalle:
                        cursor.execute("""
                            UPDATE inventarios SET cant_actual = cant_actual + ?
                            WHERE id_producto = ? AND id_espacio = ?
                        """, (cantidad, detalle["id_producto"], detalle["id_espacio"]))
                    
                    conn.commit()
                    messagebox.showinfo("Exito", f"Devolucion procesada!\n{cantidad} unidad(es) devuelta(s) y reintegrada(s) al inventario")
                    dialog.destroy()
                    form.destroy()
                except Exception as e:
                    conn.rollback()
                    messagebox.showerror("Error", str(e))
                finally:
                    conn.close()
            
            tk.Button(inner, text="Confirmar Devolucion", font=("Segoe UI", 12, "bold"),
                     bg="#dc2626", fg="white", relief="flat", cursor="hand2", command=confirmar_devolucion, pady=10).pack(fill="x", pady=(10, 0))
        
        btn_frame = tk.Frame(form, bg="white")
        btn_frame.pack(fill="x", padx=20, pady=10)
        
        tk.Button(btn_frame, text="Ver Historial", font=("Segoe UI", 10, "bold"), bg="#64748b",
                 fg="white", relief="flat", cursor="hand2", command=self.mostrar_historial_devoluciones).pack(side="right", padx=5)
        tk.Button(btn_frame, text="Procesar Devolucion", font=("Segoe UI", 11, "bold"), bg="#dc2626",
                 fg="white", relief="flat", cursor="hand2", command=procesar_devolucion).pack(side="right", padx=5)
        
    def mostrar_historial_devoluciones(self):
        form = tk.Toplevel(self.root)
        form.title("Historial de Devoluciones")
        form.geometry("700x400")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        
        header = tk.Frame(form, bg="#e74c3c")
        header.pack(fill="x")
        tk.Label(header, text="HISTORIAL DE DEVOLUCIONES", font=("Segoe UI", 14, "bold"),
                bg="#e74c3c", fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=15, pady=15)
        
        columns = ("ID", "Venta", "Fecha", "Cantidad", "Motivo", "Estado")
        tree = ttk.Treeview(card, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        tree.pack(fill="both", expand=True)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM devoluciones ORDER BY fecha_devolucion DESC")
        for row in cursor.fetchall():
            tree.insert("", "end", values=(row["id_devolucion"], row["id_venta"], row["fecha_devolucion"][:10], row["cantidad"], row["motivo"], row["estado"]))
        conn.close()
        
    def mostrar_historial_ventas(self):
        form = tk.Toplevel(self.root)
        form.title("Historial de Ventas")
        form.geometry("800x500")
        form.configure(bg=self.S.COLORS["bg_main"])
        form.transient(self.root)
        
        header = tk.Frame(form, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        tk.Label(header, text="HISTORIAL DE VENTAS", font=("Segoe UI", 14, "bold"),
                bg=self.S.COLORS["primary"], fg="white").pack(pady=12, padx=20)
        
        card = tk.Frame(form, bg="white")
        card.pack(fill="both", expand=True, padx=15, pady=15)
        
        columns = ("#", "Fecha", "Total", "Metodo", "Estado", "Vendedor")
        tree = ttk.Treeview(card, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col != "Vendedor" else 150)
        tree.pack(fill="both", expand=True)
        
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT v.*, u.nombre as vendedor FROM ventas v
            JOIN usuarios u ON v.id_usuario = u.id_usuario
            ORDER BY v.fecha_venta DESC LIMIT 100
        """)
        for row in cursor.fetchall():
            tree.insert("", "end", values=(row["id_venta"], row["fecha_venta"][:16], row["total"], row["metodo_pago"], row["estado"], row["vendedor"]))
        conn.close()
            
    def show_configuracion(self):
        self.create_toolbar([])
        
        config = get_config_sistema()
        monedas = get_monedas()
        
        if not config:
            messagebox.showerror("Error", "No se pudo cargar la configuracion")
            return
        
        main = tk.Frame(self.content_frame, bg=self.S.COLORS["bg_main"])
        main.pack(fill="both", expand=True, padx=30, pady=30)
        
        title = tk.Label(main, text="Configuracion del Sistema", font=("Segoe UI", 20, "bold"),
                        bg=self.S.COLORS["bg_main"], fg=self.S.COLORS["text_dark"])
        title.pack(anchor="w", pady=(0, 20))
        
        card = tk.Frame(main, bg="white", relief="flat", bd=0)
        card.pack(fill="x", pady=(0, 20))
        
        header = tk.Frame(card, bg=self.S.COLORS["primary"])
        header.pack(fill="x")
        tk.Label(header, text="Tienda y Moneda", font=("Segoe UI", 14, "bold"), 
                bg=self.S.COLORS["primary"], fg="white", anchor="w").pack(fill="x", pady=12, padx=20)
        
        body = tk.Frame(card, bg="white")
        body.pack(fill="x", padx=30, pady=25)
        
        row1 = tk.Frame(body, bg="white")
        row1.pack(fill="x", pady=10)
        tk.Label(row1, text="Nombre de la Tienda:", font=("Segoe UI", 12), bg="white", width=20, anchor="w").pack(side="left")
        entry_tienda = ttk.Entry(row1, width=35, font=("Segoe UI", 12))
        entry_tienda.pack(side="left", padx=10)
        entry_tienda.insert(0, config.get("nombre_tienda", "Mi Tienda"))
        
        row2 = tk.Frame(body, bg="white")
        row2.pack(fill="x", pady=10)
        tk.Label(row2, text="Moneda Local:", font=("Segoe UI", 12), bg="white", width=20, anchor="w").pack(side="left")
        var_local = tk.StringVar(value=config.get("moneda_local", "HNL"))
        combo_local = ttk.Combobox(row2, textvariable=var_local, values=[f"{m['codigo']} - {m['nombre']}" for m in monedas], width=33, state="readonly", font=("Segoe UI", 12))
        combo_local.pack(side="left", padx=10)
        
        row3 = tk.Frame(body, bg="white")
        row3.pack(fill="x", pady=10)
        tk.Label(row3, text="Moneda Extranjera:", font=("Segoe UI", 12), bg="white", width=20, anchor="w").pack(side="left")
        var_extranjera = tk.StringVar(value=config.get("moneda_extranjera", "USD"))
        combo_extranjera = ttk.Combobox(row3, textvariable=var_extranjera, values=[f"{m['codigo']} - {m['nombre']}" for m in monedas], width=33, state="readonly", font=("Segoe UI", 12))
        combo_extranjera.pack(side="left", padx=10)
        
        row4 = tk.Frame(body, bg="white")
        row4.pack(fill="x", pady=10)
        tk.Label(row4, text="Tasa de Cambio:", font=("Segoe UI", 12), bg="white", width=20, anchor="w").pack(side="left")
        entry_tasa = ttk.Entry(row4, width=15, font=("Segoe UI", 12))
        entry_tasa.pack(side="left", padx=10)
        entry_tasa.insert(0, str(config.get("tasa_cambio", 24.50)))
        
        label_preview = tk.Label(row4, text="", font=("Segoe UI", 12, "bold"), bg="white", fg=self.S.COLORS["accent"])
        label_preview.pack(side="left", padx=20)
        
        def actualizar_preview(*args):
            try:
                local = var_local.get().split(" - ")[0]
                ext = var_extranjera.get().split(" - ")[0]
                tasa_val = entry_tasa.get()
                label_preview.config(text=f"1 {ext} = {tasa_val} {local}")
            except:
                pass
        
        entry_tasa.bind("<KeyRelease>", actualizar_preview)
        var_local.trace("w", actualizar_preview)
        var_extranjera.trace("w", actualizar_preview)
        actualizar_preview()
        
        row5 = tk.Frame(body, bg="white")
        row5.pack(fill="x", pady=10)
        tk.Label(row5, text="Actualizado:", font=("Segoe UI", 11), bg="white", width=20, anchor="w").pack(side="left")
        tk.Label(row5, text=config.get("fecha_tasa", "Nunca"), font=("Segoe UI", 11, "bold"), bg="white", fg=self.S.COLORS["text_light"]).pack(side="left", padx=10)
        
        def guardar_config():
            try:
                moneda_local = var_local.get().split(" - ")[0]
                moneda_extranjera = var_extranjera.get().split(" - ")[0]
                tasa = float(entry_tasa.get())
                
                if tasa <= 0:
                    messagebox.showerror("Error", "La tasa debe ser mayor a 0")
                    return
                
                update_config_sistema(moneda_local, moneda_extranjera, tasa)
                update_config_sistema(nombre_tienda=entry_tienda.get().strip())
                
                self.root.title(f"{entry_tienda.get().strip()} - Panel de Control")
                messagebox.showinfo("Guardado", "Configuracion guardada correctamente")
                self.show_configuracion()
            except ValueError:
                messagebox.showerror("Error", "La tasa debe ser un numero")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        
        btn_guardar = tk.Button(card, text="Guardar Cambios", font=("Segoe UI", 13, "bold"),
                               bg=self.S.COLORS["success"], fg="white", relief="flat",
                               cursor="hand2", padx=40, pady=12, command=guardar_config)
        btn_guardar.pack(pady=20)
        
        card2 = tk.Frame(main, bg="white", relief="flat", bd=0)
        card2.pack(fill="x")
        
        header2 = tk.Frame(card2, bg=self.S.COLORS["accent"])
        header2.pack(fill="x")
        tk.Label(header2, text="Monedas Disponibles en el Sistema", font=("Segoe UI", 14, "bold"), 
                bg=self.S.COLORS["accent"], fg="white", anchor="w").pack(fill="x", pady=12, padx=20)
        
        body2 = tk.Frame(card2, bg="white")
        body2.pack(fill="x", padx=20, pady=20)
        
        for m in monedas:
            moneda_row = tk.Frame(body2, bg="white")
            moneda_row.pack(fill="x", pady=8)
            
            simbolo_box = tk.Frame(moneda_row, bg=self.S.COLORS["primary"], width=50, height=50)
            simbolo_box.pack(side="left", padx=(0, 15))
            simbolo_box.pack_propagate(False)
            tk.Label(simbolo_box, text=m["simbolo"], font=("Segoe UI", 20, "bold"),
                     bg=self.S.COLORS["primary"], fg="white").pack(expand=True)
            
            info_box = tk.Frame(moneda_row, bg="white")
            info_box.pack(side="left")
            tk.Label(info_box, text=f"{m['codigo']}", font=("Segoe UI", 14, "bold"),
                    bg="white", anchor="w").pack(anchor="w")
            tk.Label(info_box, text=f"{m['nombre']}", font=("Segoe UI", 10),
                    bg="white", fg="gray", anchor="w").pack(anchor="w")
            tk.Label(info_box, text=f"Tipo: {m['tipo'].upper()}", font=("Segoe UI", 9),
                    bg="white", fg=self.S.COLORS["accent"], anchor="w").pack(anchor="w")
            
    def cerrar_sesion(self):
        if messagebox.askyesno("Cerrar Sesion", "Desea cerrar la sesion?"):
            self.session.logout()
            self.root.destroy()
            
    def run(self):
        self.root.mainloop()
